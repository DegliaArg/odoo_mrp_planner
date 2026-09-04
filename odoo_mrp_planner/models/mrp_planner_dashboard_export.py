# Copyright (C) 2024 - MRP Planner
# License LGPL-3.0 or later (https://www.gnu.org/licenses/lgpl).
"""
Mixin: exportación a Excel de las tablas paginadas de los paneles de Producción
y Compras (Órdenes de Fabricación, Producido vs Programado, Órdenes de Compra,
Recepciones, Entregas y Servicios).

A diferencia del panel de Análisis de Clientes —que carga todas las filas al
navegador y exporta desde memoria— estas tablas se paginan del lado del servidor:
el frontend solo tiene la página visible. Por eso el export reutiliza los mismos
métodos de datos que alimentan cada tabla pero pidiendo el universo completo
(sin paginar), garantizando que el Excel contenga TODAS las páginas tal como
están filtradas y ordenadas en pantalla.

El armado del .xlsx (openpyxl → ir.attachment → URL de descarga) sigue el mismo
patrón que mrp_planner_dashboard_forecast_export.py.
"""
import io
import base64

from odoo import models, api

# Límite de páginas para el bucle del comparativo (única tabla con tope de
# page_size en su método de datos). 200 filas/página × 500 páginas = 100k filas,
# muy por encima de cualquier catálogo real; actúa solo como cortafuegos.
_MAX_EXPORT_PAGES = 500
_EXPORT_PAGE_SIZE = 200

# page_size "infinito" para los métodos sin tope, que devuelven todo en una llamada.
_ALL = 10 ** 9

_MO_STATE_LABEL = {
    'draft': 'Borrador', 'confirmed': 'Confirmada', 'progress': 'En progreso',
    'to_close': 'Por cerrar', 'done': 'Completada', 'cancel': 'Cancelada',
}
_REQ_STATE_LABEL = {
    'draft': 'Borrador', 'calculated': 'Calculada', 'confirmed': 'OFs creadas',
    'done': 'Hecho', 'cancel': 'Cancelada',
}
_PO_STATE_LABEL = {
    'draft': 'Borrador', 'sent': 'Enviada', 'to approve': 'Por aprobar',
    'purchase': 'Compra', 'done': 'Bloqueada', 'cancel': 'Cancelada',
}


class MrpPlannerDashboardExport(models.TransientModel):
    _inherit = 'mrp.planner.dashboard'

    # ── Builder genérico ─────────────────────────────────────────────────────

    @api.model
    def _planner_build_xlsx(self, sheet_name, filename, columns, rows):
        """
        Arma un .xlsx a partir de columnas + filas y retorna la URL de descarga.

        :param sheet_name: str — nombre de la hoja.
        :param filename: str — nombre del archivo (con extensión .xlsx). Se usa
            también como prefijo para limpiar exports previos del mismo usuario.
        :param columns: list[dict] — cada columna con:
            - 'header' (str): encabezado.
            - 'key' (str): clave del dict de fila.
            - 'number' (bool, opcional): si True se escribe como número.
        :param rows: list[dict] — filas de datos.
        :returns: dict con 'url' (str) o 'error' (str) si openpyxl no está.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return {'error': 'openpyxl no disponible'}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (sheet_name or 'Datos')[:31]  # Excel limita el nombre de hoja a 31 chars

        hdr_fill = PatternFill('solid', fgColor='1F497D')  # Azul corporativo (igual que forecast)
        hdr_font = Font(bold=True, color='FFFFFF')

        # Encabezados
        for ci, col in enumerate(columns, start=1):
            c = ws.cell(1, ci, col.get('header', ''))
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center')

        # Datos
        for ri, row in enumerate(rows, start=2):
            for ci, col in enumerate(columns, start=1):
                val = row.get(col['key'])
                if col.get('number'):
                    ws.cell(ri, ci, val if isinstance(val, (int, float)) else 0)
                else:
                    ws.cell(ri, ci, '' if val is None else str(val))

        # Anchos: encabezado vs. contenido, con tope para nombres largos
        for ci, col in enumerate(columns, start=1):
            header_len = len(str(col.get('header', '')))
            body_len = max((len(str(r.get(col['key']) or '')) for r in rows), default=0)
            width = min(max(header_len, body_len) + 2, 45)
            ws.column_dimensions[ws.cell(1, ci).column_letter].width = max(width, 10)

        ws.freeze_panes = 'A2'  # Encabezado fijo al hacer scroll

        buf = io.BytesIO()
        wb.save(buf)
        content = base64.b64encode(buf.getvalue()).decode()

        # Limpieza de exports previos del mismo usuario (mismo prefijo). Estos
        # adjuntos no están vinculados a ningún registro, por lo que el GC de
        # Odoo nunca los borraría y se acumularían indefinidamente.
        prefix = filename.rsplit('.', 1)[0].split('_')[0]
        self.env['ir.attachment'].sudo().search([
            ('res_model', '=', False),
            ('create_uid', '=', self.env.uid),
            ('name', '=like', f'{prefix}\\_%.xlsx'),
        ]).unlink()
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': content,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {'url': f'/web/content/{attachment.id}?download=true'}

    # ── Export: Órdenes de Fabricación ───────────────────────────────────────

    @api.model
    def get_mo_ofs_export(self, date_from, date_to, warehouse_id=None,
                          sort_field=None, sort_dir='asc', search=None):
        """Exporta la tabla de OFs completa (todas las páginas) tal como está filtrada."""
        data = self.get_mo_widget_data(
            date_from, date_to, warehouse_id, sort_field, sort_dir, 1, _ALL, None, search)
        rows = []
        for m in data['mos']:
            rows.append({
                'name':             m['name'],
                'product':          m['product'],
                'qty':              m['qty'],
                'date_finished':    m['date_finished'],
                'state':            _MO_STATE_LABEL.get(m['state'], m['state']),
                'pending_delivery': m['pending_delivery'],
                'delayed':          'Sí' if m['delayed'] else 'No',
                'reschedule':       'Sí' if m['reschedule'] else 'No',
            })
        columns = [
            {'header': 'Referencia',        'key': 'name'},
            {'header': 'Producto',          'key': 'product'},
            {'header': 'Cantidad',          'key': 'qty',              'number': True},
            {'header': 'Fin planificado',   'key': 'date_finished'},
            {'header': 'Estado',            'key': 'state'},
            {'header': 'Entregas pendientes', 'key': 'pending_delivery', 'number': True},
            {'header': 'Atrasada',          'key': 'delayed'},
            {'header': 'Requiere reprogramación', 'key': 'reschedule'},
        ]
        return self._planner_build_xlsx(
            'Órdenes de fabricación', f'of_{date_from}_{date_to}.xlsx', columns, rows)

    # ── Export: Producido vs Programado ──────────────────────────────────────

    @api.model
    def get_mo_comparison_export(self, date_from, date_to, warehouse_id=None,
                                 sort_field=None, sort_dir='desc', search=None):
        """Exporta el comparativo producido vs programado completo (todas las páginas).

        Único caso que pagina en bucle: get_comparison_data topea page_size a 200,
        así que se recorren las páginas hasta juntar el total.
        """
        items = []
        page = 1
        while page <= _MAX_EXPORT_PAGES:
            data = self.get_comparison_data(
                date_from, date_to, warehouse_id, page, _EXPORT_PAGE_SIZE,
                sort_field, sort_dir, search)
            items.extend(data['items'])
            if page * _EXPORT_PAGE_SIZE >= (data.get('total') or 0):
                break
            page += 1

        rows = []
        for it in items:
            pct = it.get('pct')
            rows.append({
                'product':      it['product'],
                'uom':          it.get('uom') or '',
                'planned_qty':  it['planned_qty'],
                'produced_qty': it['produced_qty'],
                'pct':          's/plan' if pct is None else pct,
            })
        columns = [
            {'header': 'Producto',      'key': 'product'},
            {'header': 'UdM',           'key': 'uom'},
            {'header': 'Programado',    'key': 'planned_qty',  'number': True},
            {'header': 'Producido',     'key': 'produced_qty', 'number': True},
            {'header': '% Cumplimiento', 'key': 'pct'},
        ]
        return self._planner_build_xlsx(
            'Producido vs Programado', f'comparativo_{date_from}_{date_to}.xlsx', columns, rows)

    # ── Export: Órdenes de Compra / Recepciones / Entregas / Servicios ───────

    @api.model
    def get_po_export(self, tab='all', oc_filter='overdue', list_tab=None,
                      date_from=None, date_to=None, sort_field=None,
                      sort_dir='asc', search=None):
        """Exporta la lista activa del panel de compras (todas las páginas).

        La lista y las columnas se eligen según la subpestaña de movimientos
        (list_tab: 'receipts' | 'deliveries' | 'services') o, en modo OC, según
        el filtro de estado (oc_filter), replicando exactamente lo que ve el usuario.
        """
        data = self.get_po_dashboard_data(
            tab, date_from, date_to, sort_field, sort_dir, 1, _ALL, search)

        suffix = f'{date_from}_{date_to}' if date_from and date_to else 'export'

        if list_tab in ('receipts', 'deliveries'):
            picks = data['receipts'] if list_tab == 'receipts' else data['deliveries']
            rows = [{
                'name':             p['name'],
                'po_name':          p['po_name'],
                'finished_product': p.get('finished_product') or '—',
                'partner':          p['partner'],
                'scheduled_date':   p['scheduled_date'],
                'days_late':        p.get('days_late') or 0,
                'availability':     p.get('availability_label') or '—',
            } for p in picks]
            if list_tab == 'receipts':
                columns = [
                    {'header': 'Referencia',    'key': 'name'},
                    {'header': 'OC',            'key': 'po_name'},
                    {'header': 'Proveedor',     'key': 'partner'},
                    {'header': 'Fecha prevista', 'key': 'scheduled_date'},
                    {'header': 'Días de atraso', 'key': 'days_late', 'number': True},
                    {'header': 'Disponibilidad', 'key': 'availability'},
                ]
                return self._planner_build_xlsx(
                    'Recepciones', f'recepciones_{suffix}.xlsx', columns, rows)
            columns = [
                {'header': 'Referencia',       'key': 'name'},
                {'header': 'OC',               'key': 'po_name'},
                {'header': 'Prod. terminado',  'key': 'finished_product'},
                {'header': 'Proveedor',        'key': 'partner'},
                {'header': 'Fecha prevista',   'key': 'scheduled_date'},
                {'header': 'Días de atraso',   'key': 'days_late', 'number': True},
                {'header': 'Disponibilidad',   'key': 'availability'},
            ]
            return self._planner_build_xlsx(
                'Entregas', f'entregas_{suffix}.xlsx', columns, rows)

        # Modo OC (o Servicios): listas de purchase.order
        if list_tab == 'services':
            pos = data['services']
            sheet, fname = 'Servicios', f'servicios_{suffix}.xlsx'
        else:
            _by_filter = {
                'all':     ('all_pos',     'oc_aprobadas'),
                'pending': ('pending_pos', 'oc_a_tiempo'),
                'rfqs':    ('rfqs',        'oc_cotizaciones'),
                'approve': ('to_approve',  'oc_por_aprobar'),
                'overdue': ('overdue',     'oc_vencidas'),
            }
            key, base = _by_filter.get(oc_filter, _by_filter['overdue'])
            pos = data[key]
            sheet, fname = 'Órdenes de compra', f'{base}_{suffix}.xlsx'

        rows = [{
            'name':           p['name'],
            'partner':        p['partner'],
            'date_planned':   p['date_planned'],
            'amount_total':   p['amount_total'],
            'state':          _PO_STATE_LABEL.get(p.get('state'), p.get('state') or ''),
            'is_subcontract': 'Sí' if p.get('is_subcontract') else 'No',
        } for p in pos]
        columns = [
            {'header': 'Referencia',       'key': 'name'},
            {'header': 'Proveedor',        'key': 'partner'},
            {'header': 'Entrega estimada', 'key': 'date_planned'},
            {'header': 'Total',            'key': 'amount_total', 'number': True},
            {'header': 'Estado',           'key': 'state'},
            {'header': 'Subcontratación',  'key': 'is_subcontract'},
        ]
        return self._planner_build_xlsx(sheet, fname, columns, rows)
