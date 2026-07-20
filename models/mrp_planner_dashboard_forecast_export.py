# Copyright (C) 2024 - MRP Planner
# License LGPL-3.0 or later (https://www.gnu.org/licenses/lgpl).
"""
Mixin: generación del archivo Excel de exportación de forecast.
Separado de mrp_planner_dashboard_forecast.py para mantener ese archivo
dentro de un tamaño manejable.
"""
import io
import base64

from odoo import models, api


class MrpPlannerDashboardForecastExport(models.TransientModel):
    _inherit = 'mrp.planner.dashboard'

    @api.model
    def get_forecast_export(self, period_from, period_to, warehouse_ids=None):
        """
        Genera un archivo Excel (.xlsx) con el resumen de forecast vs. ÓFs y retorna la URL de descarga.

        Internamente llama a get_forecast_dashboard_data para obtener los datos y luego
        construye un libro openpyxl con:
        - Fila 1: encabezados de meses (celdas combinadas Forecast + OFs).
        - Fila 2: sub-encabezados 'Forecast' / 'OFs' por mes.
        - Filas de datos: una por producto, coloreadas según cobertura (ok/warn/critical).
        - Fila de totales al final.
        El archivo se guarda como ir.attachment y se retorna su URL de descarga.

        :param period_from:  str — mes de inicio en formato 'YYYY-MM' o 'YYYY-MM-DD'.
        :param period_to:    str — mes de fin en formato 'YYYY-MM' o 'YYYY-MM-DD'.
        :param warehouse_ids: list[int] | None — se pasa directamente a get_forecast_dashboard_data.
        :returns: dict con clave 'url' (str) apuntando al endpoint /web/content/<id>?download=true,
                  o dict con clave 'error' (str) si openpyxl no está instalado.
        :raises: cualquier excepción de openpyxl o de escritura en ir.attachment se propaga.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return {'error': 'openpyxl no disponible'}

        data = self.get_forecast_dashboard_data(period_from, period_to, warehouse_ids)
        months = data['months']
        rows   = data['rows']

        MONTHS_ES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
        def _label(ym):
            y, m = ym.split('-')
            return f"{MONTHS_ES[int(m)-1]} {y}"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Forecast'

        hdr_fill = PatternFill('solid', fgColor='1F497D')   # Azul corporativo para encabezados
        hdr_font = Font(bold=True, color='FFFFFF')
        ok_fill   = PatternFill('solid', fgColor='C6EFCE')  # Verde: cobertura >= 100 %
        warn_fill = PatternFill('solid', fgColor='FFEB9C')  # Amarillo: cobertura entre warning_pct y 100 %
        crit_fill = PatternFill('solid', fgColor='FFC7CE')  # Rojo: cobertura por debajo del umbral de alerta

        warning_pct  = data['warning_pct']

        # Fila 1: encabezados de meses (agrupados de a 2)
        col = 2
        ws.cell(1, 1, 'Artículo').font = hdr_font
        ws.cell(1, 1).fill = hdr_fill
        for ym in months:
            c1 = ws.cell(1, col, _label(ym))
            c1.font = hdr_font
            c1.fill = hdr_fill
            c1.alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
            col += 2
        ws.cell(1, col, 'Total Forecast').font = hdr_font
        ws.cell(1, col).fill = hdr_fill
        ws.cell(1, col + 1, 'Total OFs').font = hdr_font
        ws.cell(1, col + 1).fill = hdr_fill

        # Fila 2: sub-encabezados Forecast / OFs
        ws.cell(2, 1, 'Artículo').font = Font(bold=True)
        col = 2
        for _ in months:
            ws.cell(2, col, 'Forecast').font = Font(bold=True)
            ws.cell(2, col + 1, 'OFs').font = Font(bold=True)
            col += 2
        ws.cell(2, col, 'Forecast').font = Font(bold=True)
        ws.cell(2, col + 1, 'OFs').font = Font(bold=True)

        # Datos
        for r, row in enumerate(rows, start=3):
            ws.cell(r, 1, row['product'])
            col = 2
            for ci, ym in enumerate(months):
                cell = row['cells'][ci]
                fc_cell = ws.cell(r, col, cell['forecast'])
                mo_cell = ws.cell(r, col + 1, cell['mos'])
                if cell['forecast'] > 0:
                    fill = ok_fill if cell['pct'] >= 100 else (warn_fill if cell['pct'] >= warning_pct else crit_fill)
                    fc_cell.fill = fill
                    mo_cell.fill = fill
                col += 2
            ws.cell(r, col, row['total_forecast'])
            ws.cell(r, col + 1, row['total_mos'])

        # Fila de totales
        trow = len(rows) + 3
        ws.cell(trow, 1, 'TOTAL').font = Font(bold=True)
        col = 2
        for mt in data['month_totals']:
            ws.cell(trow, col, mt['forecast']).font = Font(bold=True)
            ws.cell(trow, col + 1, mt['mos']).font = Font(bold=True)
            col += 2
        ws.cell(trow, col, data['kpis']['total_forecast']).font = Font(bold=True)
        ws.cell(trow, col + 1, data['kpis']['total_mos']).font = Font(bold=True)

        ws.column_dimensions['A'].width = 30   # Ancho fijo para la columna de nombre de artículo
        for i in range(2, col + 2):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = 12  # Ancho estándar para columnas numéricas

        buf = io.BytesIO()
        wb.save(buf)
        content = base64.b64encode(buf.getvalue()).decode()

        attachment = self.env['ir.attachment'].create({
            'name': f'forecast_{period_from}_{period_to}.xlsx',
            'type': 'binary',
            'datas': content,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {'url': f'/web/content/{attachment.id}?download=true'}
