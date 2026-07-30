import base64
import io
import json
from datetime import date

from odoo import models, fields, api, _
from odoo.exceptions import AccessError, UserError

_MONTHS_ES = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
               'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']


class MrpForecastImportWizard(models.TransientModel):
    _name = 'mrp.forecast.import.wizard'
    _description = 'Importar forecast desde Excel'

    file_data = fields.Binary(string='Archivo Excel', required=True, attachment=False)
    file_name = fields.Char(string='Nombre del archivo')
    sheet_name = fields.Char(string='Hoja a importar')
    sheet_names_hint = fields.Char(string='Hojas disponibles', readonly=True,
                                   help='JSON con la lista de hojas del Excel subido.')
    company_id = fields.Many2one(
        'res.company',
        string='Empresa',
        required=True,
        default=lambda self: self.env.company,
    )

    # Resultado de la importación
    result_message = fields.Text(string='Resultado', readonly=True)
    result_html = fields.Html(compute='_compute_result_html', sanitize=True)
    state = fields.Selection([
        ('upload', 'Cargar archivo'),
        ('done', 'Resultado'),
    ], default='upload')

    @api.depends('result_message', 'state')
    def _compute_result_html(self):
        from markupsafe import Markup, escape
        for rec in self:
            if rec.state != 'done' or not rec.result_message:
                rec.result_html = Markup('')
                continue
            msg = escape(rec.result_message)
            if 'Error' in rec.result_message:
                rec.result_html = Markup(
                    '<div class="alert alert-warning d-flex align-items-start">'
                    '<i class="fa fa-exclamation-triangle me-2 mt-1"></i>'
                    f'<span style="white-space:pre-wrap">{msg}</span>'
                    '</div>'
                )
            else:
                rec.result_html = Markup(
                    '<div class="alert alert-success d-flex align-items-start">'
                    '<i class="fa fa-check-circle me-2 mt-1"></i>'
                    f'<span style="white-space:pre-wrap">{msg}</span>'
                    '</div>'
                )

    @api.onchange('file_data')
    def _onchange_file_data(self):
        if not self.file_data:
            self.sheet_names_hint = False
            self.sheet_name = False
            return
        try:
            import openpyxl
        except ImportError:
            return
        try:
            raw = base64.b64decode(self.file_data)
            wb = openpyxl.load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()
            self.sheet_names_hint = json.dumps(sheets)
            if not self.sheet_name or self.sheet_name not in sheets:
                self.sheet_name = sheets[0] if sheets else False
        except Exception:
            self.sheet_names_hint = False

    def action_import(self):
        self.ensure_one()
        # Solo Ventas-Administrador o Administrador del módulo pueden importar masivamente.
        # group_sales tiene CRUD en mrp.forecast.line; group_admin es el rol superior.
        if not (self.env.user.has_group('odoo_mrp_planner.group_sales') or
                self.env.user.has_group('odoo_mrp_planner.group_admin')):
            raise UserError(_(
                'Solo usuarios con perfil "Ventas - Administrador" o '
                '"Administrador" del módulo pueden importar forecasts masivos.'
            ))
        try:
            import openpyxl
        except ImportError:
            raise UserError('openpyxl no está disponible. Contacte al administrador.')

        if not self.file_data:
            raise UserError('Seleccioná un archivo Excel.')

        raw = base64.b64decode(self.file_data)
        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(raw), data_only=True)
        except Exception as e:
            raise UserError(f'No se pudo leer el archivo: {e}')

        if self.sheet_name and self.sheet_name in wb.sheetnames:
            ws = wb[self.sheet_name]
        else:
            ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        Forecast = self.env['mrp.forecast.line']
        Product  = self.env['product.product']
        Wh       = self.env['stock.warehouse']

        errors   = []
        to_create = []
        # Pares (product_id, period, warehouse_id, company_id) a reemplazar
        to_delete_keys = []

        for i, row in enumerate(rows, start=2):
            if not row or all(v is None for v in row):
                continue

            # Columnas: A=artículo, B=período YYYY-MM, C=cantidad, D=depósito (opc.)
            raw_prod   = str(row[0]).strip() if row[0] is not None else ''
            raw_period = str(row[1]).strip() if row[1] is not None else ''
            raw_qty    = row[2]
            raw_wh     = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''

            if not raw_prod:
                errors.append(f'Fila {i}: artículo vacío.')
                continue

            # Buscar producto por referencia interna o nombre
            product = Product.search(
                ['|', ('default_code', '=', raw_prod), ('name', '=', raw_prod),
                 ('sale_ok', '=', True)], limit=1
            )
            if not product:
                errors.append(f'Fila {i}: artículo "{raw_prod}" no encontrado o no vendible.')
                continue

            # Parsear período
            period_date = self._parse_period(raw_period)
            if not period_date:
                errors.append(f'Fila {i}: período "{raw_period}" inválido (use YYYY-MM).')
                continue

            # Cantidad
            try:
                qty = float(raw_qty)
            except (TypeError, ValueError):
                errors.append(f'Fila {i}: cantidad inválida.')
                continue

            # Depósito (opcional)
            warehouse = self.env['stock.warehouse']
            if raw_wh:
                warehouse = Wh.search(['|', ('code', '=', raw_wh), ('name', '=', raw_wh)], limit=1)
                if not warehouse:
                    errors.append(f'Fila {i}: depósito "{raw_wh}" no encontrado (se omite).')

            key = (product.id, period_date, warehouse.id or False, self.company_id.id)
            if key not in to_delete_keys:
                to_delete_keys.append(key)

            to_create.append({
                'product_id':  product.id,
                'period':      period_date,
                'forecast_qty': qty,
                'warehouse_id': warehouse.id or False,
                'company_id':  self.company_id.id,
            })

        if errors and not to_create:
            self.result_message = 'No se importó ningún registro.\n\nErrores:\n' + '\n'.join(errors)
            self.state = 'done'
            return self._reopen()

        # Eliminar registros existentes para las claves importadas
        for pid, period_d, wh_id, cid in to_delete_keys:
            domain = [
                ('product_id', '=', pid),
                ('period', '=', period_d),
                ('company_id', '=', cid),
            ]
            if wh_id:
                domain.append(('warehouse_id', '=', wh_id))
            else:
                domain.append(('warehouse_id', '=', False))
            Forecast.search(domain).unlink()

        Forecast.create(to_create)

        msg_parts = [f'{len(to_create)} línea(s) importada(s) correctamente.']
        if errors:
            msg_parts.append(f'\n{len(errors)} advertencia(s):\n' + '\n'.join(errors))
        self.result_message = '\n'.join(msg_parts)
        self.state = 'done'
        return self._reopen()

    def _reopen(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    @staticmethod
    def _parse_period(value):
        """Acepta 'YYYY-MM', 'YYYY/MM' o un datetime serial de Excel."""
        if not value:
            return None
        value = str(value).strip()
        for sep in ('-', '/'):
            if sep in value:
                parts = value.split(sep)
                if len(parts) >= 2:
                    try:
                        y, m = int(parts[0]), int(parts[1])
                        if 1 <= m <= 12 and 2000 <= y <= 2100:
                            return date(y, m, 1)
                    except (ValueError, TypeError):
                        pass
        # Intentar como número serial de Excel (raramente, cuando Excel convierte la celda)
        try:
            n = int(float(value))
            # Excel epoch: 1 = 1900-01-01, con el bug del año bisiesto 1900
            from datetime import timedelta
            epoch = date(1899, 12, 30)
            d = epoch + timedelta(days=n)
            return date(d.year, d.month, 1)
        except (ValueError, TypeError):
            pass
        return None

    @api.model
    def action_download_template(self):
        """Genera y descarga una plantilla Excel vacía."""
        if not self.env.user.has_group('odoo_mrp_planner.group_sales') and \
                not self.env.user.has_group('odoo_mrp_planner.group_admin'):
            raise AccessError(_("No tiene permisos para descargar la plantilla"))
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise UserError('openpyxl no está disponible.')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Forecast'

        headers = ['Artículo (ref. interna o nombre)', 'Período (YYYY-MM)', 'Cantidad', 'Depósito (opcional)']
        header_fill = PatternFill('solid', fgColor='1F497D')
        header_font = Font(bold=True, color='FFFFFF')

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 22

        # Fila de ejemplo
        ws.append(['PROD001', '2025-07', 100, ''])

        buf = io.BytesIO()
        wb.save(buf)
        content = base64.b64encode(buf.getvalue()).decode()

        attachment = self.env['ir.attachment'].create({
            'name': 'plantilla_forecast.xlsx',
            'type': 'binary',
            'datas': content,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }
