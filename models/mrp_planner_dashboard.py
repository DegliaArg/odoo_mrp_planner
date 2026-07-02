import logging
import pytz
from datetime import datetime

from odoo import models, fields, api, _
from odoo.addons.odoo_mrp_planner.models.mrp_schedule_mixin import no_subcontract_domain

_logger = logging.getLogger(__name__)


class MrpPlannerDashboard(models.TransientModel):
    _name = 'mrp.planner.dashboard'
    _description = 'Panel del planificador de producción'

    name = fields.Char(default='Panel del Planificador')

    # ── Alertas — contadores ─────────────────────────────────────────────────

    alert_total           = fields.Integer(compute='_compute_alert_stats')
    alert_critical        = fields.Integer(compute='_compute_alert_stats')
    alert_warning         = fields.Integer(compute='_compute_alert_stats')
    alert_mo_delayed      = fields.Integer(compute='_compute_alert_stats')
    alert_mo_upcoming     = fields.Integer(compute='_compute_alert_stats')
    alert_po_delayed      = fields.Integer(compute='_compute_alert_stats')
    alert_po_upcoming     = fields.Integer(compute='_compute_alert_stats')
    alert_po_cancelled    = fields.Integer(compute='_compute_alert_stats')
    alert_receipt_delayed = fields.Integer(compute='_compute_alert_stats')
    alert_qty_mismatch    = fields.Integer(compute='_compute_alert_stats')
    alert_mo_cancelled    = fields.Integer(compute='_compute_alert_stats')

    # ── Permisos de usuario ──────────────────────────────────────────────────

    can_see_alerts       = fields.Boolean(compute='_compute_user_permissions')
    can_see_mo           = fields.Boolean(compute='_compute_user_permissions')
    can_see_po           = fields.Boolean(compute='_compute_user_permissions')
    can_see_stock_breaks = fields.Boolean(compute='_compute_user_permissions')
    can_see_forecast     = fields.Boolean(compute='_compute_user_permissions')
    can_schedule         = fields.Boolean(compute='_compute_user_permissions')
    can_reschedule       = fields.Boolean(compute='_compute_user_permissions')
    can_edit_forecast    = fields.Boolean(compute='_compute_user_permissions')

    # ── Alertas — lista inline ───────────────────────────────────────────────

    urgent_alert_ids = fields.Many2many(
        'mrp.reschedule.alert',
        compute='_compute_inline_alerts',
        string='Alertas críticas',
    )

    # ── OFs — contadores ─────────────────────────────────────────────────────

    mo_total             = fields.Integer(compute='_compute_mo_stats')
    mo_in_progress       = fields.Integer(compute='_compute_mo_stats')
    mo_done              = fields.Integer(compute='_compute_mo_stats')
    mo_delayed           = fields.Integer(compute='_compute_mo_stats')
    mo_reschedule_needed = fields.Integer(compute='_compute_mo_stats')

    # ── OFs — listas inline ──────────────────────────────────────────────────

    delayed_mo_ids    = fields.Many2many('mrp.production', compute='_compute_inline_mos',
                                         string='OFs atrasadas')
    reschedule_mo_ids = fields.Many2many('mrp.production', compute='_compute_inline_mos',
                                         string='OFs para reprogramar')

    # ── OCs — contadores ─────────────────────────────────────────────────────

    po_rfq              = fields.Integer(compute='_compute_po_stats')
    po_to_approve       = fields.Integer(compute='_compute_po_stats')
    po_total            = fields.Integer(compute='_compute_po_stats')
    po_pending          = fields.Integer(compute='_compute_po_stats')
    po_overdue          = fields.Integer(compute='_compute_po_stats')
    po_overdue_critical = fields.Integer(compute='_compute_po_stats')

    # ── OCs — listas inline ──────────────────────────────────────────────────

    rfq_ids = fields.Many2many(
        'purchase.order',
        compute='_compute_inline_pos',
        string='Solicitudes de cotización',
    )
    to_approve_ids = fields.Many2many(
        'purchase.order',
        compute='_compute_inline_pos',
        string='Por aprobar',
    )
    overdue_po_ids = fields.Many2many(
        'purchase.order',
        compute='_compute_inline_pos',
        string='OCs vencidas',
    )

    # ── Programaciones — contadores ──────────────────────────────────────────

    request_active            = fields.Integer(compute='_compute_request_stats')
    request_calculated        = fields.Integer(compute='_compute_request_stats')
    request_reschedule_needed = fields.Integer(compute='_compute_request_stats')
    req_mos_total             = fields.Integer(compute='_compute_request_stats')
    req_mos_delayed           = fields.Integer(compute='_compute_request_stats')
    req_mos_done              = fields.Integer(compute='_compute_request_stats')

    # ── Programaciones — lista inline ────────────────────────────────────────

    active_request_ids = fields.Many2many(
        'mrp.production.request',
        compute='_compute_inline_requests',
        string='Programaciones activas',
    )

    # ── Cómputos ─────────────────────────────────────────────────────────────

    @api.depends()
    def _compute_alert_stats(self):
        Alert = self.env['mrp.reschedule.alert']
        base = [('resolved', '=', False)]
        sc_loc_ids = self.env['stock.location'].search(
            [('is_subcontracting_location', '=', True)]
        ).ids
        sc_mo_ids = self.env['mrp.production'].search(
            [('location_src_id', 'in', sc_loc_ids)]
        ).ids if sc_loc_ids else []
        # Incluir alertas sin OF (recepciones, OCs); excluir solo las de OFs SBC
        no_sc = ['|', ('production_id', '=', False),
                 ('production_id', 'not in', sc_mo_ids)] if sc_mo_ids else []
        for rec in self:
            rec.alert_total           = Alert.search_count(base + no_sc)
            rec.alert_critical        = Alert.search_count(base + no_sc + [('severity', '=', 'critical')])
            rec.alert_warning         = Alert.search_count(base + no_sc + [('severity', '=', 'warning')])
            rec.alert_mo_delayed      = Alert.search_count(base + no_sc + [('alert_type', '=', 'mo_delayed')])
            rec.alert_mo_upcoming     = Alert.search_count(base + no_sc + [('alert_type', '=', 'mo_upcoming')])
            rec.alert_po_delayed      = Alert.search_count(base + [('alert_type', '=', 'po_delayed')])
            rec.alert_po_upcoming     = Alert.search_count(base + [('alert_type', '=', 'po_upcoming')])
            rec.alert_po_cancelled    = Alert.search_count(base + [('alert_type', '=', 'po_cancelled')])
            rec.alert_receipt_delayed = Alert.search_count(base + [
                ('alert_type', '=', 'receipt_delayed'),
                ('picking_id.purchase_id', '!=', False),
                ('picking_id.return_id', '=', False),
            ])
            rec.alert_qty_mismatch    = Alert.search_count(base + no_sc + [('alert_type', '=', 'qty_mismatch')])
            rec.alert_mo_cancelled    = Alert.search_count(base + no_sc + [('alert_type', '=', 'mo_cancelled')])

    @api.depends()
    def _compute_user_permissions(self):
        u = self.env.user
        is_admin    = u.has_group('odoo_mrp_planner.group_admin') or u.has_group('base.group_system')
        has_prod_r  = u.has_group('odoo_mrp_planner.group_prod_read')
        has_prod    = u.has_group('odoo_mrp_planner.group_prod')
        has_pur     = u.has_group('odoo_mrp_planner.group_purchase')
        has_sales_r = u.has_group('odoo_mrp_planner.group_sales_read')
        has_sales   = u.has_group('odoo_mrp_planner.group_sales')
        # Sin ningún grupo del módulo → mínimo = prod lectura (no schedule, no compras, no ventas)
        no_groups = not any([is_admin, has_prod_r, has_prod, has_pur, has_sales_r, has_sales])
        can_prod  = is_admin or has_prod_r or has_prod or no_groups
        can_pur   = is_admin or has_pur
        can_sales = is_admin or has_sales_r or has_sales
        for rec in self:
            rec.can_see_alerts       = can_prod or can_pur
            rec.can_see_mo           = can_prod
            rec.can_see_po           = can_pur
            rec.can_see_stock_breaks = can_prod
            rec.can_see_forecast     = can_sales
            rec.can_schedule         = is_admin or has_prod
            rec.can_reschedule       = is_admin or has_prod
            rec.can_edit_forecast    = is_admin or has_sales

    @api.depends()
    def _compute_inline_alerts(self):
        for rec in self:
            rec.urgent_alert_ids = self.env['mrp.reschedule.alert'].search(
                [('resolved', '=', False), ('severity', '=', 'critical')],
                order='id desc',
                limit=8,
            )

    @api.depends()
    def _compute_mo_stats(self):
        MO = self.env['mrp.production']
        now = fields.Datetime.now()
        no_sc = no_subcontract_domain(self.env)
        for rec in self:
            active = MO.search([('state', 'not in', ('done', 'cancel', 'draft'))] + no_sc)
            rec.mo_total             = len(active)
            rec.mo_in_progress       = len(active.filtered(lambda m: m.state in ('progress', 'to_close')))
            rec.mo_done              = MO.search_count([('state', '=', 'done')] + no_sc)
            rec.mo_delayed           = len(active.filtered(
                lambda m: m.date_finished and m.date_finished < now
            ))
            rec.mo_reschedule_needed = len(active.filtered(lambda m: m.x_reschedule_needed))

    @api.depends()
    def _compute_inline_mos(self):
        MO = self.env['mrp.production']
        now = fields.Datetime.now()
        no_sc = no_subcontract_domain(self.env)
        for rec in self:
            rec.delayed_mo_ids = MO.search([
                ('state', 'in', ('confirmed', 'progress', 'to_close')),
                ('date_finished', '<', now),
                ('date_finished', '!=', False),
            ] + no_sc, order='date_finished asc', limit=4)
            rec.reschedule_mo_ids = MO.search([
                ('state', 'not in', ('done', 'cancel')),
                ('x_reschedule_needed', '=', True),
            ] + no_sc, order='date_start asc', limit=4)

    @api.depends()
    def _compute_po_stats(self):
        PO = self.env['purchase.order']
        now = fields.Datetime.now()
        for rec in self:
            rec.po_rfq        = PO.search_count([('state', 'in', ('draft', 'sent'))])
            rec.po_to_approve = PO.search_count([('state', '=', 'to approve')])
            # Approved (purchase + done), not fully received
            active = PO.search([('state', 'in', ('purchase', 'done')), ('receipt_status', '!=', 'full')])
            overdue = active.filtered(lambda p: p.date_planned and p.date_planned < now)
            rec.po_total            = len(active)
            rec.po_pending          = len(active.filtered(
                lambda p: not p.date_planned or p.date_planned >= now
            ))
            rec.po_overdue          = len(overdue)
            cfg = self.env['mrp.reschedule.config'].search([], limit=1)
            crit_days = cfg.alert_po_critical_days if cfg else 5
            rec.po_overdue_critical = len(overdue.filtered(
                lambda p: (now - p.date_planned).days >= crit_days
            ))

    @api.depends()
    def _compute_inline_pos(self):
        PO = self.env['purchase.order']
        now = fields.Datetime.now()
        for rec in self:
            rec.rfq_ids = PO.search([
                ('state', 'in', ('draft', 'sent')),
            ], order='date_planned asc', limit=4)
            rec.to_approve_ids = PO.search([
                ('state', '=', 'to approve'),
            ], order='date_planned asc', limit=3)
            rec.overdue_po_ids = PO.search([
                ('state', 'in', ('purchase', 'done')),
                ('date_planned', '<', now),
                ('receipt_status', 'not in', ['full']),
            ], order='date_planned asc', limit=5)

    @api.depends()
    def _compute_request_stats(self):
        Req = self.env['mrp.production.request']
        now = fields.Datetime.now()
        for rec in self:
            confirmed = Req.search([('state', '=', 'confirmed')])
            calculated = Req.search([('state', '=', 'calculated')])
            all_mos = confirmed.mapped('item_ids.production_id').filtered(lambda m: m.id)
            rec.request_active     = len(confirmed)
            rec.request_calculated = len(calculated)
            rec.request_reschedule_needed = len(confirmed.filtered(
                lambda r: any(
                    it.production_id and it.production_id.x_reschedule_needed
                    for it in r.item_ids
                )
            ))
            rec.req_mos_total   = len(all_mos)
            rec.req_mos_done    = len(all_mos.filtered(lambda m: m.state == 'done'))
            rec.req_mos_delayed = len(all_mos.filtered(
                lambda m: m.state not in ('done', 'cancel')
                and m.date_finished and m.date_finished < now
            ))

    @api.depends()
    def _compute_inline_requests(self):
        for rec in self:
            rec.active_request_ids = self.env['mrp.production.request'].search([
                ('state', 'in', ('calculated', 'confirmed')),
            ], order='id desc', limit=6)

    # ── Apertura ─────────────────────────────────────────────────────────────

    @api.model
    def action_open(self):
        rec = self.create({})
        return {
            'type': 'ir.actions.act_window',
            'name': _('Panel del planificador'),
            'res_model': 'mrp.planner.dashboard',
            'res_id': rec.id,
            'view_mode': 'form',
            'view_id': self.env.ref('odoo_mrp_planner.mrp_planner_dashboard_form').id,
            'target': 'main',
            'flags': {'withControlPanel': False},
        }

    @api.model
    def action_open_ventas(self):
        rec = self.create({})
        return {
            'type': 'ir.actions.act_window',
            'name': _('Forecast de Ventas'),
            'res_model': 'mrp.planner.dashboard',
            'res_id': rec.id,
            'view_mode': 'form',
            'view_id': self.env.ref('odoo_mrp_planner.mrp_ventas_dashboard_form').id,
            'target': 'main',
            'flags': {'withControlPanel': False},
        }

    @api.model
    def action_open_compras(self):
        rec = self.create({})
        return {
            'type': 'ir.actions.act_window',
            'name': _('Panel de Compras'),
            'res_model': 'mrp.planner.dashboard',
            'res_id': rec.id,
            'view_mode': 'form',
            'view_id': self.env.ref('odoo_mrp_planner.mrp_compras_dashboard_form').id,
            'target': 'main',
            'flags': {'withControlPanel': False},
        }

    def action_refresh(self):
        self.env['mrp.reschedule.alert']._cron_check_delays()
        return self.env['mrp.planner.dashboard'].action_open()

    def action_refresh_compras(self):
        self.env['mrp.reschedule.alert']._cron_check_delays()
        return self.env['mrp.planner.dashboard'].action_open_compras()

    # ── Accesos rápidos ──────────────────────────────────────────────────────

    def action_new_request(self):
        return {
            'type': 'ir.actions.act_window',
            'name': _('Nueva programación'),
            'res_model': 'mrp.production.request',
            'view_mode': 'form',
            'target': 'current',
        }

    def action_new_plan(self):
        return {
            'type': 'ir.actions.act_window',
            'name': _('Nuevo plan de reprogramación'),
            'res_model': 'mrp.reschedule.plan',
            'view_mode': 'form',
            'target': 'current',
        }

    # ── Navegación — alertas ─────────────────────────────────────────────────

    def _open_alerts(self, extra_domain=None):
        # Alertas sin OF (recepciones, OCs) siempre se incluyen; solo se excluyen
        # las alertas de OFs de subcontratación (production_id != False y ubicación SBC)
        no_sc = ['|', ('production_id', '=', False),
                 ('production_id.location_src_id.is_subcontracting_location', '!=', True)]
        domain = [('resolved', '=', False)] + no_sc + (extra_domain or [])
        return {
            'type': 'ir.actions.act_window',
            'name': _('Alertas'),
            'res_model': 'mrp.reschedule.alert',
            'view_mode': 'list,form',
            'domain': domain,
            'target': 'current',
        }

    def action_view_all_alerts(self):
        return self._open_alerts()

    def action_view_critical(self):
        return self._open_alerts([('severity', '=', 'critical')])

    def action_view_warning(self):
        return self._open_alerts([('severity', '=', 'warning')])

    def action_view_mo_delayed_alerts(self):
        return self._open_alerts([('alert_type', '=', 'mo_delayed')])

    def action_view_po_delayed_alerts(self):
        return self._open_alerts([('alert_type', '=', 'po_delayed')])

    def action_view_po_cancelled_alerts(self):
        return self._open_alerts([('alert_type', '=', 'po_cancelled')])

    def action_view_receipt_alerts(self):
        return self._open_alerts([
            ('alert_type', '=', 'receipt_delayed'),
            ('picking_id.purchase_id', '!=', False),
            ('picking_id.return_id', '=', False),
        ])

    def action_view_qty_mismatch_alerts(self):
        return self._open_alerts([('alert_type', '=', 'qty_mismatch')])

    def action_view_mo_cancelled_alerts(self):
        return self._open_alerts([('alert_type', '=', 'mo_cancelled')])

    def action_view_mo_upcoming_alerts(self):
        return self._open_alerts([('alert_type', '=', 'mo_upcoming')])

    def action_view_po_upcoming_alerts(self):
        return self._open_alerts([('alert_type', '=', 'po_upcoming')])

    # ── Navegación — OFs ─────────────────────────────────────────────────────

    def _open_mos(self, domain, name):
        return {
            'type': 'ir.actions.act_window',
            'name': name,
            'res_model': 'mrp.production',
            'view_mode': 'list,form',
            'domain': domain,
            'target': 'current',
        }

    def action_view_all_mos(self):
        no_sc = no_subcontract_domain(self.env)
        return self._open_mos(
            [('state', 'not in', ('done', 'cancel'))] + no_sc,
            _('OFs activas'),
        )

    def action_view_in_progress_mos(self):
        no_sc = no_subcontract_domain(self.env)
        return self._open_mos(
            [('state', 'in', ('progress', 'to_close'))] + no_sc,
            _('OFs en progreso'),
        )

    def action_view_delayed_mos(self):
        now = fields.Datetime.now()
        no_sc = no_subcontract_domain(self.env)
        return self._open_mos(
            [
                ('state', 'in', ('confirmed', 'progress', 'to_close')),
                ('date_finished', '<', now),
                ('date_finished', '!=', False),
            ] + no_sc,
            _('OFs atrasadas'),
        )

    def action_view_reschedule_needed(self):
        no_sc = no_subcontract_domain(self.env)
        return self._open_mos(
            [
                ('state', 'not in', ('done', 'cancel')),
                ('x_reschedule_needed', '=', True),
            ] + no_sc,
            _('OFs para reprogramar'),
        )

    def action_view_done_mos(self):
        no_sc = no_subcontract_domain(self.env)
        return self._open_mos(
            [('state', '=', 'done')] + no_sc,
            _('OFs completadas'),
        )

    # ── Navegación — OCs ─────────────────────────────────────────────────────

    def action_view_rfqs(self):
        return {
            'type': 'ir.actions.act_window',
            'name': _('Solicitudes de cotización'),
            'res_model': 'purchase.order',
            'view_mode': 'list,form',
            'domain': [('state', 'in', ('draft', 'sent'))],
            'target': 'current',
        }

    def action_view_to_approve(self):
        return {
            'type': 'ir.actions.act_window',
            'name': _('Por aprobar'),
            'res_model': 'purchase.order',
            'view_mode': 'list,form',
            'domain': [('state', '=', 'to approve')],
            'target': 'current',
        }

    def action_view_pending_pos(self):
        now = fields.Datetime.now()
        return {
            'type': 'ir.actions.act_window',
            'name': _('OCs a tiempo'),
            'res_model': 'purchase.order',
            'view_mode': 'list,form',
            'domain': [
                ('state', 'in', ('purchase', 'done')),
                '|', ('date_planned', '>=', now), ('date_planned', '=', False),
            ],
            'target': 'current',
        }

    def action_view_overdue_pos(self):
        now = fields.Datetime.now()
        return {
            'type': 'ir.actions.act_window',
            'name': _('OCs vencidas'),
            'res_model': 'purchase.order',
            'view_mode': 'list,form',
            'domain': [
                ('state', 'in', ('purchase', 'done')),
                ('date_planned', '<', now),
                ('receipt_status', 'not in', ['full']),
            ],
            'target': 'current',
        }

    def action_view_all_pos(self):
        return {
            'type': 'ir.actions.act_window',
            'name': _('Órdenes de compra aprobadas'),
            'res_model': 'purchase.order',
            'view_mode': 'list,form',
            'domain': [('state', 'in', ('purchase', 'done'))],
            'target': 'current',
        }

    # ── Navegación — Programaciones ──────────────────────────────────────────

    def action_view_active_requests(self):
        return {
            'type': 'ir.actions.act_window',
            'name': _('Programaciones con OFs creadas'),
            'res_model': 'mrp.production.request',
            'view_mode': 'list,form',
            'domain': [('state', '=', 'confirmed')],
            'target': 'current',
        }

    def action_view_calculated_requests(self):
        return {
            'type': 'ir.actions.act_window',
            'name': _('Programaciones calculadas'),
            'res_model': 'mrp.production.request',
            'view_mode': 'list,form',
            'domain': [('state', '=', 'calculated')],
            'target': 'current',
        }

    def action_view_requests_reschedule(self):
        return {
            'type': 'ir.actions.act_window',
            'name': _('Programaciones con reprogramación pendiente'),
            'res_model': 'mrp.production.request',
            'view_mode': 'list,form',
            'domain': [
                ('state', '=', 'confirmed'),
                ('item_ids.production_id.x_reschedule_needed', '=', True),
            ],
            'target': 'current',
        }

    # ── Filtros de sector (WC tags) — usados por widgets de OFs ────────────

    @api.model
    def get_wc_tags(self):
        """Tags de centros de trabajo que tienen al menos un WC activo."""
        Tag = self.env['mrp.workcenter.tag']
        Wc  = self.env['mrp.workcenter']
        result = []
        for tag in Tag.search([]):
            if Wc.search_count([('tag_ids', 'in', tag.id), ('active', '=', True)]):
                result.append({'id': tag.id, 'name': tag.name})
        return result

    @api.model
    def get_wc_chart_data(self, date_from, date_to, tag_id=None):
        """Carga de CTs: horas disponibles vs ejecutado/pendiente/libre por centro de trabajo."""
        first_day = datetime.strptime(date_from, '%Y-%m-%d')
        last_day  = datetime.strptime(date_to,   '%Y-%m-%d').replace(hour=23, minute=59, second=59)

        domain = [('active', '=', True)]
        if tag_id:
            domain.append(('tag_ids', 'in', int(tag_id)))
        workcenters = self.env['mrp.workcenter'].search(domain)

        labels, avail_list = [], []
        ejecutado_list, pendiente_list, tiempo_muerto_list = [], [], []

        def _avail_hours(calendar, dt_start, dt_end, efficiency):
            try:
                h = calendar.get_work_hours_count(
                    dt_start.replace(tzinfo=pytz.UTC),
                    dt_end.replace(tzinfo=pytz.UTC),
                    compute_leaves=False,
                )
                return h * (efficiency or 100.0) / 100.0
            except Exception as e:
                _logger.debug("WC chart: error calendario %s: %s", calendar.name, e)
                weekly = sum(
                    a.hour_to - a.hour_from
                    for a in calendar.attendance_ids
                    if not a.date_from and not a.date_to
                )
                span = (dt_end - dt_start).days + 1
                return weekly * (span / 7.0) * (efficiency or 100.0) / 100.0

        def _overlap_hours(wo, range_start, range_end):
            start = wo.date_start
            end   = wo.date_finished
            if not start:
                return 0.0
            ov_start = max(start, range_start)
            ov_end   = min(end, range_end) if end else range_end
            if ov_start >= ov_end:
                return 0.0
            if not end:
                return (wo.duration_expected or 0.0) / 60.0
            total_secs = (end - start).total_seconds()
            if total_secs <= 0:
                return (wo.duration_expected or 0.0) / 60.0
            proportion = (ov_end - ov_start).total_seconds() / total_secs
            return (wo.duration_expected or 0.0) / 60.0 * proportion

        for wc in workcenters:
            efficiency = wc.time_efficiency or 100.0
            avail = 0.0
            if wc.resource_calendar_id:
                avail = _avail_hours(wc.resource_calendar_id, first_day, last_day, efficiency)

            wos = self.env['mrp.workorder'].search([
                ('workcenter_id', '=', wc.id),
                ('state', 'not in', ('cancel',)),
                ('date_start', '!=', False),
                ('date_start', '<=', fields.Datetime.to_string(last_day)),
                '|',
                ('date_finished', '>=', fields.Datetime.to_string(first_day)),
                ('date_finished', '=', False),
                ('production_id.location_src_id.is_subcontracting_location', '!=', True),
            ])

            ejecutado = sum(
                _overlap_hours(w, first_day, last_day)
                for w in wos if w.state == 'done'
            )
            pendiente = sum(
                _overlap_hours(w, first_day, last_day)
                for w in wos if w.state not in ('done', 'cancel')
            )
            tiempo_muerto = max(0.0, avail - ejecutado - pendiente)

            if avail == 0.0 and ejecutado == 0.0 and pendiente == 0.0:
                continue

            labels.append(wc.name)
            avail_list.append(round(avail, 1))
            ejecutado_list.append(round(ejecutado, 1))
            pendiente_list.append(round(pendiente, 1))
            tiempo_muerto_list.append(round(tiempo_muerto, 1))

        tot_avail = sum(avail_list)
        tot_ejec  = sum(ejecutado_list)
        tot_pend  = sum(pendiente_list)
        tot_plan  = tot_ejec + tot_pend
        tot_libre = sum(tiempo_muerto_list)
        carga_pct = round(tot_plan / tot_avail * 100, 1) if tot_avail > 0 else 0.0

        return {
            'labels':          labels,
            'available_hours': avail_list,
            'ejecutado':       ejecutado_list,
            'pendiente':       pendiente_list,
            'tiempo_muerto':   tiempo_muerto_list,
            'totals': {
                'disponible':   round(tot_avail, 1),
                'planificado':  round(tot_plan,  1),
                'carga_pct':    carga_pct,
                'ejecutado':    round(tot_ejec,  1),
                'pendiente':    round(tot_pend,  1),
                'tiempo_libre': round(tot_libre, 1),
            },
        }

    # ── Widget OCs con pestañas ──────────────────────────────────────────────

    @api.model
    def get_po_dashboard_data(self, filter_type='all', date_from=None, date_to=None, sort_field=None, sort_dir='asc', page=1, page_size=50):
        """Datos de OCs filtrados por tipo y rango de fecha de entrega."""
        PO      = self.env['purchase.order']
        Picking = self.env['stock.picking']
        now     = fields.Datetime.now()

        _sd = 'desc' if sort_dir == 'desc' else 'asc'
        _rev = (_sd == 'desc')
        _PO_FIELD = {
            'name': 'name', 'partner': 'partner_id',
            'date_planned': 'date_planned', 'amount_total': 'amount_total',
        }
        _PICK_FIELD = {
            'name': 'name', 'partner': 'partner_id',
            'scheduled_date': 'scheduled_date', 'overdue': 'scheduled_date',
            'availability': 'state',
        }
        po_f       = _PO_FIELD.get(sort_field, 'date_planned')
        pick_f     = _PICK_FIELD.get(sort_field, 'scheduled_date')
        po_order   = f'{po_f} {_sd}'
        pick_order = f'{pick_f} {_sd}'
        offset     = (max(1, page) - 1) * page_size

        sc_domain = []
        if filter_type == 'purchase':
            sc_domain = [('subcontract_production_ids', '=', False)]
        elif filter_type == 'subcontract':
            sc_domain = [('subcontract_production_ids', '!=', False)]

        date_domain = []
        if date_from:
            date_domain.append(('date_order', '>=', date_from + ' 00:00:00'))
        if date_to:
            date_domain.append(('date_order', '<=', date_to + ' 23:59:59'))

        sched_domain = []
        if date_from:
            sched_domain.append(('scheduled_date', '>=', date_from + ' 00:00:00'))
        if date_to:
            sched_domain.append(('scheduled_date', '<=', date_to + ' 23:59:59'))

        rfq_dom      = [('state', 'in', ('draft', 'sent'))] + sc_domain + date_domain
        approve_dom  = [('state', '=', 'to approve')] + sc_domain + date_domain
        approved_dom = [('state', 'in', ('purchase', 'done'))] + sc_domain + date_domain

        approved = PO.search(approved_dom)
        overdue  = approved.filtered(lambda p: p.date_planned and p.date_planned < now)
        pending  = approved.filtered(lambda p: not p.date_planned or p.date_planned >= now)

        # Leer umbral crítico OC desde config
        cfg = self.env['mrp.reschedule.config'].search([], limit=1)
        po_crit_days = cfg.alert_po_critical_days if cfg else 5

        def _po_dict(po):
            return {
                'id':               po.id,
                'name':             po.name,
                'partner':          po.partner_id.display_name if po.partner_id else '',
                'date_planned':     po.date_planned.strftime('%d/%m/%Y') if po.date_planned else '—',
                'amount_total':     po.amount_total,
                'is_subcontract':   bool(po.subcontract_production_ids),
            }

        def _move_qty(m):
            """En Odoo 18, 'quantity' es el campo unificado (antes quantity_done
            en move_lines). Para pickings de subcontratación, reserved_availability=0
            pero quantity=demanda. Usamos el máximo de ambos para cubrir ambos casos."""
            return max(
                getattr(m, 'quantity', 0) or 0,
                getattr(m, 'reserved_availability', 0) or 0,
            )

        def _pick_avail(p):
            """En Odoo 16+, 'partially_available' fue eliminado como estado.
            Los pickings parcialmente reservados quedan en 'assigned'.
            Detectamos la diferencia comparando qty disponible vs demanda."""
            if p.state == 'assigned':
                is_partial = any(
                    _move_qty(m) < m.product_uom_qty - 0.001
                    for m in p.move_ids if m.state not in ('done', 'cancel')
                )
                return 'partially_available' if is_partial else 'assigned'
            if p.state == 'confirmed':
                has_any = any(
                    _move_qty(m) > 0.001
                    for m in p.move_ids if m.state not in ('done', 'cancel')
                )
                return 'partially_available' if has_any else 'confirmed'
            return p.state

        _AVAIL_LABEL = {
            'assigned':            'Disponible',
            'partially_available': 'Parcialmente',
            'confirmed':           'No disponible',
            'waiting':             'No disponible',
        }

        _has_raw_mo  = 'raw_material_production_id' in self.env['stock.move']._fields
        _has_po_line = 'purchase_line_id' in self.env['mrp.production']._fields

        def _trace_mo(moves, depth=0):
            """Traza move_dest_ids recursivamente buscando una OF de subcontratación."""
            if depth > 10:
                return None
            for mv in moves:
                if _has_raw_mo and mv.raw_material_production_id:
                    return mv.raw_material_production_id
                if mv.production_id:
                    return mv.production_id
                result = _trace_mo(mv.move_dest_ids, depth + 1)
                if result:
                    return result
            return None

        def _delivery_mo_s1(p):
            """Estrategia 1: campo directo raw_material_production_id (usa prefetch)."""
            if not _has_raw_mo:
                return None
            for mv in p.move_ids:
                if mv.raw_material_production_id:
                    return mv.raw_material_production_id
            return None

        def _delivery_info(p):
            """Para un picking de entrega devuelve (po_name, finished_product).
            Busca la OF de subcontratación por 4 estrategias y extrae ambos datos."""
            # Estrategia 1: campo directo en el move (usa prefetch, O(1))
            mo = _delivery_mo_s1(p)
            # Estrategia 2: trazar move_dest_ids
            if not mo:
                mo = _trace_mo(p.move_ids)
            # Estrategia 3: desde líneas de la OC
            if not mo and p.purchase_id:
                for line in p.purchase_id.order_line:
                    for mv in line.move_ids:
                        if _has_raw_mo and mv.raw_material_production_id:
                            mo = mv.raw_material_production_id
                            break
                    if mo:
                        break
            # Estrategia 4: por grupo de abastecimiento (fallback)
            if not mo and p.group_id:
                mo = self.env['mrp.production'].search(
                    [('procurement_group_id', '=', p.group_id.id)], limit=1
                )

            finished = (mo.product_id.display_name if mo and mo.product_id else None) or '—'

            po_name = '—'
            if mo and _has_po_line:
                try:
                    if mo.purchase_line_id and mo.purchase_line_id.order_id:
                        po_name = mo.purchase_line_id.order_id.name
                except Exception:
                    pass

            return po_name, finished

        def _pick_dict(p, include_lines=False):
            avail = _pick_avail(p)
            is_incoming = p.picking_type_code == 'incoming'
            if is_incoming:
                po_name  = p.purchase_id.name if p.purchase_id else '—'
                finished = None
            else:
                po_name, finished = _delivery_info(p)
            result = {
                'id':               p.id,
                'name':             p.name,
                'po_name':          po_name,
                'finished_product': finished,
                'partner':          p.partner_id.display_name if p.partner_id else '',
                'scheduled_date': p.scheduled_date.strftime('%d/%m/%Y') if p.scheduled_date else '—',
                'state':          p.state,
                'overdue':        bool(p.scheduled_date and p.scheduled_date < now),
                'days_late':      max(0, (now - p.scheduled_date).days) if p.scheduled_date and p.scheduled_date < now else 0,
                'availability':   avail,
                'availability_label': _AVAIL_LABEL.get(avail, '—'),
                'lines':          [],
                'is_incoming':    is_incoming,
            }
            if include_lines:
                result['lines'] = [{
                    'product':  m.product_id.display_name,
                    'demand':   m.product_uom_qty,
                    'reserved': (getattr(m, 'quantity_done', None) or getattr(m, 'quantity', 0) or 0)
                                if is_incoming else _move_qty(m),
                    'uom':      m.product_uom.name if m.product_uom else '',
                } for m in p.move_ids if m.product_id and m.state not in ('done', 'cancel')]
            return result

        rfqs_list       = PO.search(rfq_dom,     order=po_order)
        to_approve_list = PO.search(approve_dom, order=po_order)

        # ── Separar servicios ────────────────────────────────────────────────
        show_svc = bool(cfg and cfg.show_po_services_tab)

        def _is_svc(po):
            lines = po.order_line.filtered(lambda l: l.product_id)
            return bool(lines) and all(l.product_id.type == 'service' for l in lines)

        if show_svc:
            rfqs_svc        = rfqs_list.filtered(_is_svc)
            rfqs_list       = rfqs_list - rfqs_svc
            approve_svc     = to_approve_list.filtered(_is_svc)
            to_approve_list = to_approve_list - approve_svc
            approved_svc    = approved.filtered(_is_svc)
            approved        = approved - approved_svc
            overdue         = overdue - approved_svc
            pending         = pending - approved_svc
            services_rs     = (rfqs_svc | approve_svc | approved_svc).sorted(po_f, reverse=_rev)
        else:
            services_rs     = self.env['purchase.order']

        overdue_list  = overdue.sorted(po_f,  reverse=_rev)
        all_pos_list  = approved.sorted(po_f, reverse=_rev)
        pending_list  = pending.sorted(po_f,  reverse=_rev)

        # Sort por nombre de partner (no por ID): hacerlo en Python para que
        # sea correcto en todas las páginas, no solo la primera.
        if sort_field == 'partner':
            _pk = lambda r: (r.partner_id.display_name or '').lower()
            rfqs_list       = rfqs_list.sorted(_pk,       reverse=_rev)
            to_approve_list = to_approve_list.sorted(_pk, reverse=_rev)
            overdue_list    = overdue_list.sorted(_pk,    reverse=_rev)
            all_pos_list    = all_pos_list.sorted(_pk,    reverse=_rev)
            pending_list    = pending_list.sorted(_pk,    reverse=_rev)
            services_rs     = services_rs.sorted(_pk,     reverse=_rev)



        rfqs_pg       = rfqs_list[offset:offset + page_size]
        to_approve_pg = to_approve_list[offset:offset + page_size]
        overdue_pg    = overdue_list[offset:offset + page_size]
        all_pos_pg    = all_pos_list[offset:offset + page_size]
        pending_pg    = pending_list[offset:offset + page_size]

        # ── Recepciones (incoming pickings linked to POs) ────────────────────
        receipt_sc = []
        if filter_type == 'purchase':
            receipt_sc = [('purchase_id.subcontract_production_ids', '=', False)]
        elif filter_type == 'subcontract':
            receipt_sc = [('purchase_id.subcontract_production_ids', '!=', False)]

        receipts = Picking.search([
            ('state', 'not in', ['done', 'cancel']),
            ('picking_type_code', '=', 'incoming'),
            ('purchase_id', '!=', False),
            ('return_id', '=', False),
        ] + receipt_sc, order=pick_order)

        overdue_receipts = receipts.filtered(lambda p: p.scheduled_date and p.scheduled_date < now)

        # ── Entregas (component deliveries to subcontractors) ───────────────
        # La OC no es un M2O en estos pickings, es texto. El único campo
        # confiable es el destino: ubicación de subcontratación.
        deliveries = Picking.search([
            ('state', 'not in', ['done', 'cancel']),
            ('location_dest_id.is_subcontracting_location', '=', True),
        ] + sched_domain, order=pick_order)

        # Prefetch para evitar N+1 en sort y en _pick_dict
        (receipts | deliveries).mapped('move_ids')
        if deliveries and _has_raw_mo:
            deliveries.mapped('move_ids.raw_material_production_id.product_id')
            if _has_po_line:
                deliveries.mapped('move_ids.raw_material_production_id.purchase_line_id.order_id')

        # Sort por partner/availability/po_name en pickings (Python, cross-página)
        if sort_field == 'partner':
            _ppk = lambda p: (p.partner_id.display_name or '').lower()
            receipts   = receipts.sorted(_ppk,   reverse=_rev)
            deliveries = deliveries.sorted(_ppk, reverse=_rev)
        elif sort_field == 'availability':
            _AO = {'assigned': 0, 'partially_available': 1, 'confirmed': 2, 'waiting': 3}
            _ak = lambda p: _AO.get(_pick_avail(p), 99)
            receipts   = receipts.sorted(_ak,   reverse=_rev)
            deliveries = deliveries.sorted(_ak, reverse=_rev)
        elif sort_field == 'finished_product':
            def _fpk(p):
                mo = _delivery_mo_s1(p)
                return (mo.product_id.display_name if mo and mo.product_id else '').lower()
            deliveries = deliveries.sorted(_fpk, reverse=_rev)
        elif sort_field == 'po_name':
            def _dok(p):
                mo = _delivery_mo_s1(p)
                if mo and _has_po_line:
                    try:
                        if mo.purchase_line_id and mo.purchase_line_id.order_id:
                            return mo.purchase_line_id.order_id.name.lower()
                    except Exception:
                        pass
                return ''
            receipts   = receipts.sorted(lambda p: (p.purchase_id.name or '').lower(), reverse=_rev)
            deliveries = deliveries.sorted(_dok, reverse=_rev)

        receipts_pg        = receipts[offset:offset + page_size]
        deliveries_pg      = deliveries[offset:offset + page_size]
        services_pg        = services_rs[offset:offset + page_size]
        overdue_deliveries = deliveries.filtered(lambda p: p.scheduled_date and p.scheduled_date < now)

        return {
            'kpis': {
                'rfq':              len(rfqs_list),
                'to_approve':       len(to_approve_list),
                'total':            len(approved),
                'pending':          len(pending),
                'overdue':          len(overdue),
                'overdue_critical': len(overdue.filtered(
                    lambda p: (now - p.date_planned).days >= po_crit_days
                )),
                'receipts_total':    len(receipts),
                'receipts_overdue':  len(overdue_receipts),
                'deliveries_total':  len(deliveries),
                'deliveries_overdue': len(overdue_deliveries),
                'services_total':    len(services_rs),
                'po_critical_days':  po_crit_days,
            },
            'show_services_tab': show_svc,
            'rfqs':        [_po_dict(p) for p in rfqs_pg],
            'to_approve':  [_po_dict(p) for p in to_approve_pg],
            'overdue':     [_po_dict(p) for p in overdue_pg],
            'all_pos':     [_po_dict(p) for p in all_pos_pg],
            'pending_pos': [_po_dict(p) for p in pending_pg],
            'receipts':    [_pick_dict(p, True)  for p in receipts_pg],
            'deliveries':  [_pick_dict(p, True)  for p in deliveries_pg],
            'services':    [_po_dict(p) for p in services_pg],
        }

    # ── Widget OFs filtrable ─────────────────────────────────────────────────

    @api.model
    def get_filtered_mos(self, date_from, date_to, tag_id=None):
        """OFs activas (sin subcontratación) que solapan con el rango, filtradas por sector."""
        first_day = datetime.strptime(date_from, '%Y-%m-%d')
        last_day  = datetime.strptime(date_to,   '%Y-%m-%d').replace(hour=23, minute=59, second=59)

        no_sc = no_subcontract_domain(self.env)
        domain = [
            ('state', 'not in', ('done', 'cancel')),
            ('date_start', '<=', fields.Datetime.to_string(last_day)),
            '|',
            ('date_finished', '>=', fields.Datetime.to_string(first_day)),
            '&',
            ('date_finished', '=', False),
            ('date_start', '>=', fields.Datetime.to_string(first_day)),
        ] + no_sc

        mos = self.env['mrp.production'].search(domain, order='date_finished asc')

        if tag_id:
            tag_id = int(tag_id)
            mos = mos.filtered(
                lambda m: any(
                    tag_id in w.workcenter_id.tag_ids.ids
                    for w in m.workorder_ids
                    if w.workcenter_id
                )
            )

        now = fields.Datetime.now()
        result = []
        for mo in mos:
            result.append({
                'id':            mo.id,
                'name':          mo.name,
                'product':       mo.product_id.display_name if mo.product_id else '',
                'qty':           mo.product_qty,
                'date_finished': mo.date_finished.strftime('%d/%m/%Y') if mo.date_finished else '',
                'state':         mo.state,
                'delayed':       bool(mo.date_finished and mo.date_finished < now),
                'reschedule':    bool(mo.x_reschedule_needed),
            })
        return result

    # ── Widget Alertas ───────────────────────────────────────────────────────

    @api.model
    def get_alert_stats(self, states=None):
        """Contadores de alertas de OFs filtrados por estado y sin subcontratadas."""
        Alert = self.env['mrp.reschedule.alert']
        base  = [('resolved', '=', False)]
        sc_loc_ids = self.env['stock.location'].search(
            [('is_subcontracting_location', '=', True)]
        ).ids
        sc_mo_ids = self.env['mrp.production'].search(
            [('location_src_id', 'in', sc_loc_ids)]
        ).ids if sc_loc_ids else []
        # Incluir alertas sin OF (recepciones, OCs); excluir solo las de OFs SBC
        no_sc = ['|', ('production_id', '=', False),
                 ('production_id', 'not in', sc_mo_ids)] if sc_mo_ids else []

        def cnt(alert_type):
            return Alert.search_count(base + no_sc + [('alert_type', '=', alert_type)])

        return {
            'mo_delayed':   cnt('mo_delayed'),
            'mo_upcoming':  cnt('mo_upcoming'),
            'qty_mismatch': cnt('qty_mismatch'),
            'mo_cancelled': cnt('mo_cancelled'),
            'critical':     Alert.search_count(base + no_sc + [('severity', '=', 'critical')]),
        }

    # ── Widget OFs con pestañas ──────────────────────────────────────────────

    @api.model
    def get_mo_widget_data(self, date_from, date_to, tag_id=None, sort_field=None, sort_dir='asc', page=1, page_size=50, states=None):
        """KPIs + lista de OFs activas en el rango, filtradas por sector y estados."""
        first_day = datetime.strptime(date_from, '%Y-%m-%d')
        last_day  = datetime.strptime(date_to,   '%Y-%m-%d').replace(hour=23, minute=59, second=59)

        _sd = 'desc' if sort_dir == 'desc' else 'asc'
        _MO_FIELD = {
            'name': 'name', 'product': 'product_id', 'qty': 'product_qty',
            'date_finished': 'date_finished', 'state': 'state',
            'delayed': 'date_finished', 'reschedule': 'x_reschedule_needed',
        }
        mo_f     = _MO_FIELD.get(sort_field, 'date_finished')
        mo_order = f'{mo_f} {_sd}'

        no_sc = no_subcontract_domain(self.env)
        # Estados activos seleccionados (excluye 'done' que tiene su propio dominio)
        active_states = [s for s in (states or []) if s != 'done'] if states else []
        state_clause  = [('state', 'in', active_states)] if active_states else [('state', 'not in', ('done', 'cancel'))]
        domain = state_clause + [
            ('date_start', '<=', fields.Datetime.to_string(last_day)),
            '|',
            ('date_finished', '>=', fields.Datetime.to_string(first_day)),
            '&',
            ('date_finished', '=', False),
            ('date_start', '>=', fields.Datetime.to_string(first_day)),
        ] + no_sc

        mos = self.env['mrp.production'].search(domain, order=mo_order)

        if tag_id:
            tag_id = int(tag_id)
            tag_filter = lambda m: any(
                tag_id in w.workcenter_id.tag_ids.ids
                for w in m.workorder_ids if w.workcenter_id
            )
            mos = mos.filtered(tag_filter)

        # OFs finalizadas en el mismo rango de fechas
        done_domain = [
            ('state', '=', 'done'),
            ('date_finished', '>=', fields.Datetime.to_string(first_day)),
            ('date_finished', '<=', fields.Datetime.to_string(last_day)),
        ] + no_sc
        done_mos = self.env['mrp.production'].search(done_domain)
        if tag_id:
            done_mos = done_mos.filtered(tag_filter)

        offset   = (max(1, page) - 1) * page_size
        mos_page = mos[offset:offset + page_size]

        now = fields.Datetime.now()

        # Batch-compute pending outgoing deliveries per product (ítem 7)
        product_ids = list({mo.product_id.id for mo in mos_page if mo.product_id})
        if product_ids:
            out_moves = self.env['stock.move'].search([
                ('product_id', 'in', product_ids),
                ('state', 'not in', ('done', 'cancel')),
                ('picking_id.picking_type_id.code', '=', 'outgoing'),
            ])
            pending_out = {}
            for m in out_moves:
                pid = m.product_id.id
                pending_out[pid] = pending_out.get(pid, 0.0) + m.product_uom_qty
        else:
            pending_out = {}

        def _mo_dict(mo):
            return {
                'id':               mo.id,
                'name':             mo.name,
                'product':          mo.product_id.display_name if mo.product_id else '',
                'qty':              mo.product_qty,
                'date_finished':    mo.date_finished.strftime('%d/%m/%Y') if mo.date_finished else '',
                'state':            mo.state,
                'delayed':          bool(mo.date_finished and mo.date_finished < now),
                'reschedule':       bool(mo.x_reschedule_needed),
                'pending_delivery': round(pending_out.get(mo.product_id.id, 0.0), 2),
            }

        return {
            'kpis': {
                'total':       len(mos),
                'in_progress': len(mos.filtered(lambda m: m.state in ('progress', 'to_close'))),
                'delayed':     len(mos.filtered(lambda m: m.date_finished and m.date_finished < now)),
                'reschedule':  len(mos.filtered(lambda m: m.x_reschedule_needed)),
                'done':        len(done_mos),
                'partial':     len(mos.filtered(lambda m: m.state == 'to_close')),
            },
            'mos': [_mo_dict(m) for m in mos_page],
        }

    @api.model
    def get_mo_kpi_counts(self, date_from, date_to):
        """Contadores de KPIs usando los mismos dominios que los botones de navegación JS.
        Garantiza que el número del card coincida exactamente con el listado al hacer click."""
        now   = fields.Datetime.now()
        MO    = self.env['mrp.production']
        no_sc = no_subcontract_domain(self.env)
        dFrom = date_from + ' 00:00:00'
        dTo   = date_to   + ' 23:59:59'
        date_d = [('date_finished', '>=', dFrom), ('date_finished', '<=', dTo)]
        active = [('state', 'not in', ('done', 'cancel', 'draft'))]
        now_s  = fields.Datetime.to_string(now)

        return {
            'total':       MO.search_count(active + date_d + no_sc),
            'in_progress': MO.search_count([('state', 'in', ('progress', 'to_close'))] + date_d + no_sc),
            'delayed':     MO.search_count(active + [('date_finished', '<', now_s)] + date_d + no_sc),
            'reschedule':  MO.search_count(active + [('x_reschedule_needed', '=', True)] + date_d + no_sc),
            'done':        MO.search_count([('state', '=', 'done')] + date_d + no_sc),
            'partial':     MO.search_count([('state', '=', 'to_close')] + date_d + no_sc),
        }

    @api.model
    def get_request_widget_data(self, sort_field=None, sort_dir='asc', page=1, page_size=50):
        """KPIs + lista de programaciones activas."""
        Req = self.env['mrp.production.request']
        now = fields.Datetime.now()

        _sd = 'desc' if sort_dir == 'desc' else 'asc'
        _REQ_FIELD = {'name': 'name', 'start_from': 'start_from', 'state': 'state'}
        req_f = _REQ_FIELD.get(sort_field, 'id')

        confirmed  = Req.search([('state', '=', 'confirmed')])
        calculated = Req.search([('state', '=', 'calculated')])
        all_active = (confirmed | calculated).sorted(req_f, reverse=(_sd == 'desc'))
        all_mos    = confirmed.mapped('item_ids.production_id').filtered(lambda m: m.id)

        offset          = (max(1, page) - 1) * page_size
        all_active_page = all_active[offset:offset + page_size]

        def _req_dict(r):
            mos = r.item_ids.mapped('production_id').filtered(lambda m: m.id)
            return {
                'id':          r.id,
                'name':        r.name,
                'start_from':  r.start_from.strftime('%d/%m/%Y') if r.start_from else '—',
                'state':       r.state,
                'mos_total':   len(mos),
                'mos_done':    len(mos.filtered(lambda m: m.state == 'done')),
                'mos_delayed': len(mos.filtered(
                    lambda m: m.state not in ('done', 'cancel')
                    and m.date_finished and m.date_finished < now
                )),
            }

        return {
            'kpis': {
                'total':       len(all_active),
                'active':      len(confirmed),
                'calculated':  len(calculated),
                'reschedule':  len(confirmed.filtered(
                    lambda r: any(
                        it.production_id and it.production_id.x_reschedule_needed
                        for it in r.item_ids
                    )
                )),
                'mos_delayed': len(all_mos.filtered(
                    lambda m: m.state not in ('done', 'cancel')
                    and m.date_finished and m.date_finished < now
                )),
            },
            'requests': [_req_dict(r) for r in all_active_page],
        }

    @api.model
    def get_comparison_data(self, date_from, date_to, tag_id=None):
        """Producido vs programado por producto en el rango."""
        first_day = datetime.strptime(date_from, '%Y-%m-%d')
        last_day  = datetime.strptime(date_to,   '%Y-%m-%d').replace(hour=23, minute=59, second=59)

        no_sc = no_subcontract_domain(self.env)

        done_mos = self.env['mrp.production'].search([
            ('state', '=', 'done'),
            ('date_finished', '>=', fields.Datetime.to_string(first_day)),
            ('date_finished', '<=', fields.Datetime.to_string(last_day)),
        ] + no_sc)

        active_mos = self.env['mrp.production'].search([
            ('state', 'not in', ('done', 'cancel')),
            ('date_start', '<=', fields.Datetime.to_string(last_day)),
            '|',
            ('date_finished', '>=', fields.Datetime.to_string(first_day)),
            '&',
            ('date_finished', '=', False),
            ('date_start', '>=', fields.Datetime.to_string(first_day)),
        ] + no_sc)

        all_mos = done_mos | active_mos

        if tag_id:
            tag_id = int(tag_id)
            all_mos = all_mos.filtered(
                lambda m: any(
                    tag_id in w.workcenter_id.tag_ids.ids
                    for w in m.workorder_ids if w.workcenter_id
                )
            )

        product_data = {}
        for mo in all_mos:
            pid = mo.product_id.id
            if not pid:
                continue
            if pid not in product_data:
                product_data[pid] = {
                    'product_id':   pid,
                    'product':      mo.product_id.display_name,
                    'uom':          mo.product_uom_id.name if mo.product_uom_id else '',
                    'planned_qty':  0.0,
                    'produced_qty': 0.0,
                }
            product_data[pid]['planned_qty']  += mo.product_qty
            product_data[pid]['produced_qty'] += mo.qty_produced

        items = sorted(product_data.values(), key=lambda x: x['planned_qty'], reverse=True)
        for item in items:
            item['pct'] = round(
                item['produced_qty'] / item['planned_qty'] * 100, 1
            ) if item['planned_qty'] > 0 else 0.0
            item['planned_qty']  = round(item['planned_qty'],  2)
            item['produced_qty'] = round(item['produced_qty'], 2)

        total_planned  = sum(x['planned_qty']  for x in items)
        total_produced = sum(x['produced_qty'] for x in items)
        pct = round(total_produced / total_planned * 100, 1) if total_planned > 0 else 0.0

        filtered_done = all_mos.filtered(lambda m: m.state == 'done')
        return {
            'kpis': {
                'planned':  round(total_planned,  2),
                'produced': round(total_produced, 2),
                'pct':      pct,
                'ofs_done': len(filtered_done),
            },
            'items': items[:20],
        }

    # ── Backwards compat (paneles de detalle) ─────────────────────────────────

    def action_open_mos_dashboard(self):
        return self.env['mrp.planner.detail.dashboard'].action_open_for_category('mos')

    def action_open_pos_dashboard(self):
        return self.env['mrp.planner.detail.dashboard'].action_open_for_category('pos')

    def action_open_requests_dashboard(self):
        return self.env['mrp.planner.detail.dashboard'].action_open_for_category('requests')

    # ── Widget quiebres de stock ─────────────────────────────────────────────

    @api.model
    def get_internal_locations(self):
        """Devuelve todas las ubicaciones internas activas para el selector del widget."""
        locations = self.env['stock.location'].search(
            [('usage', '=', 'internal'), ('active', '=', True)],
            order='complete_name',
        )
        return [{'id': l.id, 'name': l.complete_name} for l in locations]

    # ── Forecast ─────────────────────────────────────────────────────────────

    @api.model
    def get_warehouses_for_forecast(self):
        user = self.env.user
        if user.mrp_planner_all_warehouses or not user.mrp_planner_warehouse_ids:
            whs = self.env['stock.warehouse'].search([], order='name')
        else:
            whs = user.mrp_planner_warehouse_ids.sorted('name')
        return [{'id': w.id, 'name': w.name} for w in whs]

    @api.model
    def get_wc_load_data(self):
        """Carga de centros de trabajo: horas planificadas en OFs activas."""
        domain = [
            ('state', 'in', ['ready', 'progress', 'pending', 'confirmed']),
            ('production_id.state', 'in', ['confirmed', 'progress', 'to_close']),
            ('production_id.location_src_id.is_subcontracting_location', '!=', True),
        ]
        wc_map = {}
        for wo in self.env['mrp.workorder'].search(domain):
            wc = wo.workcenter_id
            if not wc:
                continue
            hours = (wo.duration_expected or 0.0) / 60.0
            if wc.id not in wc_map:
                wc_map[wc.id] = {'id': wc.id, 'name': wc.name, 'hours': 0.0, 'count': 0}
            wc_map[wc.id]['hours'] += hours
            wc_map[wc.id]['count'] += 1

        rows = sorted(wc_map.values(), key=lambda x: -x['hours'])[:10]
        if not rows:
            return []
        max_h = rows[0]['hours'] or 1.0
        for r in rows:
            r['hours'] = round(r['hours'], 1)
            r['bar_pct'] = min(100, round(r['hours'] / max_h * 100))
        return rows

    @api.model
    def get_forecast_dashboard_data(self, period_from, period_to, warehouse_ids=None):
        """Devuelve KPIs y tabla pivotada forecast vs ÓFs para el rango de meses indicado."""
        from datetime import date as _date
        import calendar as _calendar

        warehouse_ids = warehouse_ids or []

        def _parse_ym(ym):
            parts = ym.split('-')
            y, m = int(parts[0]), int(parts[1])
            d = int(parts[2]) if len(parts) >= 3 else 1
            return _date(y, m, d)

        def _months_between(d_from, d_to):
            months = []
            d = _date(d_from.year, d_from.month, 1)
            while d <= _date(d_to.year, d_to.month, 1):
                months.append(f"{d.year}-{d.month:02d}")
                if d.month == 12:
                    d = _date(d.year + 1, 1, 1)
                else:
                    d = _date(d.year, d.month + 1, 1)
            return months

        try:
            d_from = _parse_ym(period_from)
            d_to   = _parse_ym(period_to)
        except Exception:
            return {'kpis': {}, 'months': [], 'month_totals': [], 'rows': [],
                    'warning_pct': 70, 'critical_pct': 50}

        months = _months_between(d_from, d_to)

        cfg = self.env['mrp.reschedule.config'].search([], limit=1)
        warning_pct    = cfg.forecast_warning_pct    if cfg else 70
        critical_pct   = cfg.forecast_critical_pct   if cfg else 50
        rotation_unit  = (cfg.forecast_rotation_unit if cfg else None) or 'days'
        acc_formula    = (cfg.forecast_acc_formula   if cfg else None) or 'simple'

        # Estados de OF configurados
        mo_states = []
        if cfg:
            if cfg.forecast_mo_state_draft:     mo_states.append('draft')
            if cfg.forecast_mo_state_confirmed: mo_states.append('confirmed')
            if cfg.forecast_mo_state_progress:  mo_states.append('progress')
            if cfg.forecast_mo_state_to_close:  mo_states.append('to_close')
            if cfg.forecast_mo_state_done:      mo_states.append('done')
        if not mo_states:
            mo_states = ['confirmed', 'progress', 'to_close']

        last_day_of_to = d_to  # día exacto seleccionado por el usuario

        # ── Forecast lines ────────────────────────────────────────────────────
        fc_domain = [
            ('period', '>=', _date(d_from.year, d_from.month, 1)),
            ('period', '<=', last_day_of_to),
            ('company_id', '=', self.env.company.id),
        ]

        fc_lines = self.env['mrp.forecast.line'].search(fc_domain)

        # Estructura: {product_id: {month_str: forecast_qty}}
        fc_data = {}
        for line in fc_lines:
            pid = line.product_id.id
            ym  = f"{line.period.year}-{line.period.month:02d}"
            if pid not in fc_data:
                fc_data[pid] = {
                    'product':         line.product_id.display_name,
                    'product_tmpl_id': line.product_id.product_tmpl_id.id,
                }
            fc_data[pid][ym] = fc_data[pid].get(ym, 0.0) + line.forecast_qty

        # ── ÓFs planificadas ──────────────────────────────────────────────────
        mo_domain = [
            ('state', 'in', mo_states),
            ('date_finished', '>=', fields.Datetime.to_string(
                datetime.combine(d_from, datetime.min.time())
            )),
            ('date_finished', '<=', fields.Datetime.to_string(
                datetime.combine(last_day_of_to, datetime.max.time())
            )),
        ] + no_subcontract_domain(self.env)
        mos = self.env['mrp.production'].search(mo_domain)

        # Estructura: {product_id: {month_str: qty}}
        mo_data = {}
        for mo in mos:
            if not mo.product_id or not mo.date_finished:
                continue
            pid = mo.product_id.id
            df  = mo.date_finished
            ym  = f"{df.year}-{df.month:02d}"
            if ym not in months:
                continue
            if pid not in mo_data:
                mo_data[pid] = {}
            mo_data[pid][ym] = mo_data[pid].get(ym, 0.0) + mo.product_qty

        # ── Ids de productos con forecast ──────────────────────────────────────
        all_product_ids      = set(fc_data.keys())
        all_product_ids_list = list(all_product_ids)
        n_months             = len(months) or 1

        # ── Movimientos de salida completados (entregado) ──────────────────────
        del_line_domain = [
            ('state', '=', 'done'),
            ('picking_id.picking_type_id.code', '=', 'outgoing'),
            ('date', '>=', fields.Datetime.to_string(
                datetime.combine(d_from, datetime.min.time()))),
            ('date', '<=', fields.Datetime.to_string(
                datetime.combine(last_day_of_to, datetime.max.time()))),
            ('product_id', 'in', all_product_ids_list),
            ('company_id', '=', self.env.company.id),
        ]
        del_data = {}   # {product_id: {ym: qty}}
        for ml in self.env['stock.move.line'].search(del_line_domain):
            pid = ml.product_id.id
            dt  = ml.date
            if not dt:
                continue
            ym = f"{dt.year}-{dt.month:02d}"
            if ym not in months:
                continue
            del_data.setdefault(pid, {})
            del_data[pid][ym] = del_data[pid].get(ym, 0.0) + ml.quantity

        # ── Demanda real: pedidos de venta confirmados ─────────────────────────
        so_data = {}    # {product_id: {ym: qty}}
        try:
            so_domain = [
                ('order_id.state', 'in', ('sale', 'done')),
                ('order_id.date_order', '>=', fields.Datetime.to_string(
                    datetime.combine(d_from, datetime.min.time()))),
                ('order_id.date_order', '<=', fields.Datetime.to_string(
                    datetime.combine(last_day_of_to, datetime.max.time()))),
                ('product_id', 'in', all_product_ids_list),
                ('company_id', '=', self.env.company.id),
            ]
            for line in self.env['sale.order.line'].search(so_domain):
                pid = line.product_id.id
                if not line.order_id.date_order:
                    continue
                dt  = line.order_id.date_order
                ym  = f"{dt.year}-{dt.month:02d}"
                if ym not in months:
                    continue
                so_data.setdefault(pid, {})
                so_data[pid][ym] = so_data[pid].get(ym, 0.0) + line.product_uom_qty
        except Exception:
            pass    # módulo sale no disponible

        # ── Stock actual (snapshot) ───────────────────────────────────────────
        stock_data = {}   # {product_id: qty}
        quant_domain = [
            ('location_id.usage', '=', 'internal'),
            ('product_id', 'in', all_product_ids_list),
            ('company_id', '=', self.env.company.id),
        ]
        if warehouse_ids:
            wh_recs  = self.env['stock.warehouse'].browse(warehouse_ids)
            loc_ids  = wh_recs.mapped('lot_stock_id').ids
            if loc_ids:
                quant_domain.append(('location_id', 'in', loc_ids))
        for q in self.env['stock.quant'].search(quant_domain):
            pid = q.product_id.id
            stock_data[pid] = stock_data.get(pid, 0.0) + q.quantity

        # ── Construir filas ────────────────────────────────────────────────────
        rows = []
        # Categorías de venta por product.template
        tmpl_ids = [fc_data[pid].get('product_tmpl_id') for pid in all_product_ids
                    if fc_data[pid].get('product_tmpl_id')]
        if tmpl_ids:
            tmpl_info = {}
            for t in self.env['product.template'].browse(tmpl_ids):
                tmpl_info[t.id] = {
                    'sale_category': t.x_sale_category or '',
                    'product_categ': t.categ_id.display_name if t.categ_id else '',
                    'product_types': ', '.join(t.x_product_type_ids.mapped('name')),
                }
        else:
            tmpl_info = {}

        for pid in all_product_ids:
            pname    = fc_data[pid]['product']
            pid_del  = del_data.get(pid, {})
            pid_so   = so_data.get(pid, {})
            stock_qty = round(stock_data.get(pid, 0.0), 2)
            cells            = []
            tot_fc           = 0.0
            tot_mos          = 0.0
            tot_del          = 0.0
            tot_so           = 0.0
            _mape_acc_sum    = 0.0   # Σ precisión por período (MAPE)
            _mape_acc_count  = 0     # períodos con del > 0 (MAPE)
            _wape_abs_err    = 0.0   # Σ|error| ponderado por real (WAPE)
            _wmape_abs_err   = 0.0   # Σ|error| ponderado por forecast (WMAPE)

            for ym in months:
                fc_qty  = fc_data[pid].get(ym, 0.0)
                mo_qty  = mo_data.get(pid, {}).get(ym, 0.0)
                del_qty = pid_del.get(ym, 0.0)
                so_qty  = pid_so.get(ym, 0.0)
                pct      = round(mo_qty  / fc_qty * 100, 1) if fc_qty > 0 else 0.0
                svc_rate = round(del_qty / so_qty * 100, 1) if so_qty > 0 else None

                abs_err = abs(so_qty - fc_qty)
                # Precisión calculada vs demanda real (so_qty), no entregado
                if so_qty > 0:
                    _mape_acc_sum   += max(0.0, 100.0 - abs_err / so_qty * 100)
                    _mape_acc_count += 1
                    _wape_abs_err   += abs_err
                if fc_qty > 0:
                    _wmape_abs_err  += abs_err
                # Valor de celda solo para la fórmula configurada
                if acc_formula in ('mape', 'wape'):
                    fc_acc = round(max(0.0, 100.0 - abs_err / so_qty * 100), 1) if so_qty > 0 else None
                elif acc_formula == 'wmape':
                    fc_acc = round(max(0.0, 100.0 - abs_err / fc_qty * 100), 1) if fc_qty > 0 else None
                elif acc_formula == 'bias':
                    fc_acc = round((so_qty - fc_qty) / fc_qty * 100, 1) if fc_qty > 0 else None
                else:  # simple
                    fc_acc = round(so_qty / fc_qty * 100, 1) if fc_qty > 0 else None

                demand_gap_pct = round((so_qty - fc_qty) / fc_qty * 100, 1) if fc_qty > 0 else None

                cells.append({
                    'month':           ym,
                    'forecast':        round(fc_qty,  2),
                    'mos':             round(mo_qty,  2),
                    'pct':             pct,
                    'delivered':       round(del_qty, 2),
                    'so_demand':       round(so_qty,  2),
                    'service_rate':    svc_rate,
                    'forecast_acc':    fc_acc,
                    'demand_gap_pct':  demand_gap_pct,
                })
                tot_fc  += fc_qty
                tot_mos += mo_qty
                tot_del += del_qty
                tot_so  += so_qty

            tot_pct = round(tot_mos / tot_fc * 100, 1) if tot_fc > 0 else 0.0
            tot_svc = round(tot_del / tot_so * 100, 1) if tot_so > 0 else None
            if acc_formula == 'mape':
                tot_acc = round(_mape_acc_sum / _mape_acc_count, 1) if _mape_acc_count > 0 else None
            elif acc_formula == 'wape':
                tot_acc = round(max(0.0, 100.0 - _wape_abs_err / tot_so * 100), 1) if tot_so > 0 else None
            elif acc_formula == 'wmape':
                tot_acc = round(max(0.0, 100.0 - _wmape_abs_err / tot_fc * 100), 1) if tot_fc > 0 else None
            elif acc_formula == 'bias':
                tot_acc = round((tot_so - tot_fc) / tot_fc * 100, 1) if tot_fc > 0 else None
            else:
                tot_acc = round(tot_so / tot_fc * 100, 1) if tot_fc > 0 else None

            avg_monthly_del = tot_del / n_months
            if avg_monthly_del > 0:
                rot_months = round(stock_qty / avg_monthly_del, 1)
                rot_days   = int(round(stock_qty / avg_monthly_del * 30))
            else:
                rot_months = None
                rot_days   = None

            rows.append({
                'product_id':         pid,
                'product_tmpl_id':    fc_data[pid].get('product_tmpl_id'),
                'product':            pname,
                'cells':              cells,
                'stock_qty':          stock_qty,
                'rotation_days':      rot_days,
                'rotation_months':    rot_months,
                'total_forecast':     round(tot_fc,  2),
                'total_mos':          round(tot_mos, 2),
                'total_pct':          tot_pct,
                'total_delivered':    round(tot_del, 2),
                'total_so_demand':    round(tot_so,  2),
                'total_service_rate': tot_svc,
                'total_forecast_acc': tot_acc,
                'demand_gap_pct': round((tot_so - tot_fc) / tot_fc * 100, 1) if tot_fc > 0 else None,
                'acc_all': {
                    'simple': round(tot_so / tot_fc * 100, 1) if tot_fc > 0 else None,
                    'mape':   round(_mape_acc_sum / _mape_acc_count, 1) if _mape_acc_count > 0 else None,
                    'wape':   round(max(0.0, 100.0 - _wape_abs_err / tot_so * 100), 1) if tot_so > 0 else None,
                    'wmape':  round(max(0.0, 100.0 - _wmape_abs_err / tot_fc * 100), 1) if tot_fc > 0 else None,
                    'bias':   round((tot_so - tot_fc) / tot_fc * 100, 1) if tot_fc > 0 else None,
                },
                'sale_category':      tmpl_info.get(fc_data[pid].get('product_tmpl_id'), {}).get('sale_category', ''),
                'product_categ':      tmpl_info.get(fc_data[pid].get('product_tmpl_id'), {}).get('product_categ', ''),
                'product_types':      tmpl_info.get(fc_data[pid].get('product_tmpl_id'), {}).get('product_types', ''),
                '_mape_acc_sum':      _mape_acc_sum,
                '_mape_acc_count':    _mape_acc_count,
                '_wape_abs_err':      _wape_abs_err,
                '_wmape_abs_err':     _wmape_abs_err,
            })

        rows.sort(key=lambda r: r['product'].lower())

        # ── Totales por mes ────────────────────────────────────────────────────
        month_totals = []
        for i, ym in enumerate(months):
            mfc = sum(r['cells'][i]['forecast']  for r in rows)
            mmo = sum(r['cells'][i]['mos']       for r in rows)
            mdl = sum(r['cells'][i]['delivered'] for r in rows)
            mso = sum(r['cells'][i]['so_demand'] for r in rows)
            month_totals.append({
                'month':     ym,
                'forecast':  round(mfc, 2),
                'mos':       round(mmo, 2),
                'delivered': round(mdl, 2),
                'so_demand': round(mso, 2),
            })

        # ── KPIs globales ──────────────────────────────────────────────────────
        total_fc   = sum(r['total_forecast']  for r in rows)
        total_mos  = sum(r['total_mos']       for r in rows)
        total_del  = sum(r['total_delivered'] for r in rows)
        total_so   = sum(r['total_so_demand'] for r in rows)

        # Producción de OFs para productos SIN línea de forecast (solo vendibles)
        no_fc_mo_pids = [pid for pid in mo_data if pid not in all_product_ids]
        mos_no_fc = 0.0
        if no_fc_mo_pids:
            sale_ok_pids = set(
                self.env['product.product'].browse(no_fc_mo_pids)
                .filtered(lambda p: p.sale_ok).ids
            )
            mos_no_fc = round(sum(
                sum(v.values()) for pid, v in mo_data.items()
                if pid not in all_product_ids and pid in sale_ok_pids
            ), 2)

        # Entregado para productos SIN línea de forecast (solo vendibles)
        delivered_no_fc = 0.0
        try:
            no_fc_del_domain = [
                ('state', '=', 'done'),
                ('picking_id.picking_type_id.code', '=', 'outgoing'),
                ('date', '>=', fields.Datetime.to_string(
                    datetime.combine(d_from, datetime.min.time()))),
                ('date', '<=', fields.Datetime.to_string(
                    datetime.combine(last_day_of_to, datetime.max.time()))),
                ('product_id.sale_ok', '=', True),
                ('company_id', '=', self.env.company.id),
            ]
            if all_product_ids_list:
                no_fc_del_domain.append(('product_id', 'not in', all_product_ids_list))
            groups = self.env['stock.move.line'].read_group(no_fc_del_domain, ['quantity:sum'], [])
            delivered_no_fc = round((groups[0]['quantity'] or 0.0) if groups else 0.0, 2)
        except Exception:
            delivered_no_fc = 0.0

        # Demanda de SOs en el período para productos SIN línea de forecast
        so_demand_no_fc = 0.0
        try:
            no_fc_domain = [
                ('order_id.state', 'in', ('sale', 'done')),
                ('order_id.date_order', '>=', fields.Datetime.to_string(
                    datetime.combine(d_from, datetime.min.time()))),
                ('order_id.date_order', '<=', fields.Datetime.to_string(
                    datetime.combine(last_day_of_to, datetime.max.time()))),
                ('company_id', '=', self.env.company.id),
            ]
            if all_product_ids_list:
                no_fc_domain.append(('product_id', 'not in', all_product_ids_list))
            groups = self.env['sale.order.line'].read_group(no_fc_domain, ['product_uom_qty:sum'], [])
            so_demand_no_fc = round((groups[0]['product_uom_qty'] or 0.0) if groups else 0.0, 2)
        except Exception:
            so_demand_no_fc = 0.0

        coverage   = round(total_mos / total_fc * 100, 1) if total_fc > 0 else 0.0
        at_risk    = sum(1 for r in rows if r['total_forecast'] > 0 and r['total_pct'] < warning_pct)
        ovr_svc = round(total_del / total_so * 100, 1) if total_so > 0 else None
        if acc_formula == 'mape':
            all_mape_sum   = sum(r['_mape_acc_sum']   for r in rows)
            all_mape_count = sum(r['_mape_acc_count'] for r in rows)
            ovr_acc = round(all_mape_sum / all_mape_count, 1) if all_mape_count > 0 else None
        elif acc_formula == 'wape':
            all_wape_err = sum(r['_wape_abs_err'] for r in rows)
            ovr_acc = round(max(0.0, 100.0 - all_wape_err / total_so * 100), 1) if total_so > 0 else None
        elif acc_formula == 'wmape':
            all_wmape_err  = sum(r['_wmape_abs_err'] for r in rows)
            total_fc_wmape = sum(r['total_forecast']  for r in rows if r['total_forecast'] > 0)
            ovr_acc = round(max(0.0, 100.0 - all_wmape_err / total_fc_wmape * 100), 1) if total_fc_wmape > 0 else None
        elif acc_formula == 'bias':
            ovr_acc = round((total_so - total_fc) / total_fc * 100, 1) if total_fc > 0 else None
        else:
            ovr_acc = round(total_so / total_fc * 100, 1) if total_fc > 0 else None

        # Limpiar campos internos antes de enviar al frontend
        _internal = ('_mape_acc_sum', '_mape_acc_count', '_wape_abs_err', '_wmape_abs_err')
        for r in rows:
            for k in _internal:
                r.pop(k, None)

        return {
            'kpis': {
                'total_forecast':       round(total_fc,  2),
                'total_so_demand':      round(total_so,  2),
                'total_mos':            round(total_mos, 2),
                'total_delivered':      round(total_del, 2),
                'gap':                  round(total_mos - total_fc, 2),
                'mos_gap_pct':          round((total_mos - total_fc) / total_fc * 100, 1) if total_fc > 0 else None,
                'demand_gap':           round(total_so - total_fc, 2),
                'demand_gap_pct':       round((total_so - total_fc) / total_fc * 100, 1) if total_fc > 0 else None,
                'coverage_pct':         coverage,
                'at_risk':              at_risk,
                'total_products':       len(rows),
                'overall_service_rate': ovr_svc,
                'overall_forecast_acc': ovr_acc,
                'so_demand_no_fc':      so_demand_no_fc,
                'mos_no_fc':            mos_no_fc,
                'delivered_no_fc':      delivered_no_fc,
            },
            'months':        months,
            'month_totals':  month_totals,
            'rows':          rows,
            'warning_pct':   warning_pct,
            'critical_pct':  critical_pct,
            'rotation_unit': rotation_unit,
            'acc_formula':   acc_formula,
            'mo_states':     mo_states,
        }

    @api.model
    def get_product_mos_for_forecast(self, product_id, period_from, period_to, warehouse_ids=None):
        """OFs de un producto para el acordeón del widget de forecast."""
        from datetime import date as _date, datetime
        import calendar as _cal

        d_from = _date(int(period_from[:4]), int(period_from[5:7]), 1)
        d_to_y, d_to_m = int(period_to[:4]), int(period_to[5:7])
        last_day = _date(d_to_y, d_to_m, _cal.monthrange(d_to_y, d_to_m)[1])

        domain = [
            ('product_id', '=', product_id),
            ('state', 'not in', ['cancel']),
            ('date_finished', '>=', fields.Datetime.to_string(
                datetime.combine(d_from, datetime.min.time()))),
            ('date_finished', '<=', fields.Datetime.to_string(
                datetime.combine(last_day, datetime.max.time()))),
        ] + no_subcontract_domain(self.env)
        if warehouse_ids:
            wh_recs = self.env['stock.warehouse'].browse(warehouse_ids)
            loc_ids = wh_recs.mapped('lot_stock_id').ids
            if loc_ids:
                domain.append(('location_dest_id', 'in', loc_ids))

        mos = self.env['mrp.production'].search(domain, limit=100, order='date_finished asc')
        state_labels = {
            'draft':     'Borrador',
            'confirmed': 'Confirmada',
            'progress':  'En progreso',
            'to_close':  'Por cerrar',
            'done':      'Hecha',
            'cancel':    'Cancelada',
        }
        return [{
            'id':           mo.id,
            'name':         mo.name,
            'state':        mo.state,
            'state_label':  state_labels.get(mo.state, mo.state),
            'product_qty':  round(mo.product_qty, 2),
            'qty_produced': round(mo.qty_produced, 2),
            'uom':          mo.product_uom_id.name if mo.product_uom_id else '',
            'date_start':   mo.date_start.strftime('%Y-%m-%d')    if mo.date_start    else None,
            'date_finished': mo.date_finished.strftime('%Y-%m-%d') if mo.date_finished else None,
        } for mo in mos]

    @api.model
    def get_forecast_export(self, period_from, period_to, warehouse_ids=None):
        """Genera un Excel con el forecast y las ÓFs planificadas y retorna la URL de descarga."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return {'error': 'openpyxl no disponible'}
        import io, base64

        data = self.get_forecast_dashboard_data(period_from, period_to, warehouse_ids)
        months = data['months']
        rows   = data['rows']

        MONTHS_ES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
        def _label(ym):
            y, m = ym.split('-')
            return f"{MONTHS_ES[int(m)-1]} {y}"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Forecast'

        hdr_fill = PatternFill('solid', fgColor='1F497D')
        hdr_font = Font(bold=True, color='FFFFFF')
        ok_fill   = PatternFill('solid', fgColor='C6EFCE')
        warn_fill = PatternFill('solid', fgColor='FFEB9C')
        crit_fill = PatternFill('solid', fgColor='FFC7CE')

        warning_pct  = data['warning_pct']

        # Fila 1: encabezados de meses (agrupados de a 2)
        col = 2
        ws.cell(1, 1, 'Artículo').font = hdr_font
        ws.cell(1, 1).fill = hdr_fill
        for ym in months:
            c1 = ws.cell(1, col, _label(ym))
            c1.font = hdr_font
            c1.fill = hdr_fill
            c1.alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
            col += 2
        ws.cell(1, col, 'Total Forecast').font = hdr_font
        ws.cell(1, col).fill = hdr_fill
        ws.cell(1, col + 1, 'Total OFs').font = hdr_font
        ws.cell(1, col + 1).fill = hdr_fill

        # Fila 2: sub-encabezados Forecast / OFs
        ws.cell(2, 1, 'Artículo').font = Font(bold=True)
        col = 2
        for _ in months:
            ws.cell(2, col, 'Forecast').font = Font(bold=True)
            ws.cell(2, col + 1, 'OFs').font = Font(bold=True)
            col += 2
        ws.cell(2, col, 'Forecast').font = Font(bold=True)
        ws.cell(2, col + 1, 'OFs').font = Font(bold=True)

        # Datos
        for r, row in enumerate(rows, start=3):
            ws.cell(r, 1, row['product'])
            col = 2
            for ci, ym in enumerate(months):
                cell = row['cells'][ci]
                fc_cell = ws.cell(r, col, cell['forecast'])
                mo_cell = ws.cell(r, col + 1, cell['mos'])
                if cell['forecast'] > 0:
                    fill = ok_fill if cell['pct'] >= 100 else (warn_fill if cell['pct'] >= warning_pct else crit_fill)
                    fc_cell.fill = fill
                    mo_cell.fill = fill
                col += 2
            ws.cell(r, col, row['total_forecast'])
            ws.cell(r, col + 1, row['total_mos'])

        # Fila de totales
        trow = len(rows) + 3
        ws.cell(trow, 1, 'TOTAL').font = Font(bold=True)
        col = 2
        for mt in data['month_totals']:
            ws.cell(trow, col, mt['forecast']).font = Font(bold=True)
            ws.cell(trow, col + 1, mt['mos']).font = Font(bold=True)
            col += 2
        ws.cell(trow, col, data['kpis']['total_forecast']).font = Font(bold=True)
        ws.cell(trow, col + 1, data['kpis']['total_mos']).font = Font(bold=True)

        ws.column_dimensions['A'].width = 30
        for i in range(2, col + 2):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = 12

        buf = io.BytesIO()
        wb.save(buf)
        content = base64.b64encode(buf.getvalue()).decode()

        attachment = self.env['ir.attachment'].create({
            'name': f'forecast_{period_from}_{period_to}.xlsx',
            'type': 'binary',
            'datas': content,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {'url': f'/web/content/{attachment.id}?download=true'}

    # ── Widget quiebres de stock ─────────────────────────────────────────────

    @api.model
    def get_stock_break_data(self, filter_type='all', sort_field=None, sort_dir='asc', page=1, page_size=20, search='', location_ids=None):
        """Productos con sale_ok=True, su stock en la/las ubicación/es indicadas y el mínimo
        del punto de reorden con ruta Fabricación. location_ids puede ser una lista de IDs
        o None/[] para usar la ubicación configurada por defecto."""
        _empty_kpis = {'total': 0, 'broken': 0, 'ok': 0, 'no_min': 0}

        # Normalizar location_ids a lista de enteros
        if location_ids and isinstance(location_ids, int):
            location_ids = [location_ids]
        elif not location_ids:
            location_ids = []

        if location_ids:
            locations = self.env['stock.location'].browse(location_ids).filtered(lambda l: l.exists())
        else:
            loc_param = self.env['ir.config_parameter'].sudo().get_param(
                'mrp_reschedule.stock_location_id')
            loc_id = int(loc_param) if loc_param else False
            loc = self.env['stock.location'].browse(loc_id) if loc_id else self.env['stock.location']
            locations = loc if loc_id and loc.exists() else self.env['stock.location']

        if not locations:
            return {'error': 'no_location', 'kpis': _empty_kpis,
                    'products': [], 'location_name': '', 'location_ids': [],
                    'location_id': False, 'total_filtered': 0}

        location_name = ' + '.join(locations.mapped('complete_name'))

        # Ruta fabricación: primero por xmlid, fallback por nombre
        mfg_route = self.env.ref('mrp.route_warehouse0_manufacture', raise_if_not_found=False)
        if not mfg_route:
            mfg_route = self.env['stock.route'].search(
                [('name', 'ilike', 'manufactur')], limit=1)

        # Productos vendibles activos
        product_domain = [('sale_ok', '=', True), ('active', '=', True), ('type', '=', 'consu')]
        if search:
            product_domain += ['|', ('name', 'ilike', search), ('default_code', 'ilike', search)]
        products = self.env['product.product'].search(product_domain)
        if not products:
            return {'error': None, 'kpis': _empty_kpis,
                    'products': [], 'location_name': location_name, 'total_filtered': 0}

        product_ids = products.ids

        # Puntos de reorden con ruta fabricación → min_qty por producto
        op_domain = [('product_id', 'in', product_ids)]
        if mfg_route:
            op_domain.append(('route_id', '=', mfg_route.id))
        orderpoints = self.env['stock.warehouse.orderpoint'].search(op_domain)
        min_qty_map = {}
        forecast_map = {}
        for op in orderpoints:
            pid = op.product_id.id
            op_min = op.product_min_qty
            if pid not in min_qty_map or op_min > min_qty_map[pid]:
                min_qty_map[pid] = op_min
                forecast_map[pid] = getattr(op, 'qty_forecast', None)

        # Stock en ubicaciones seleccionadas (batch via read_group)
        quant_groups = self.env['stock.quant'].read_group(
            [('product_id', 'in', product_ids),
             ('location_id', 'child_of', locations.ids)],
            ['product_id', 'quantity:sum'],
            ['product_id'],
        )
        qty_map = {g['product_id'][0]: g['quantity'] for g in quant_groups}

        # Construir filas
        rows = []
        for p in products:
            qty     = round(qty_map.get(p.id, 0.0), 3)
            min_qty = min_qty_map.get(p.id)
            has_min = min_qty is not None
            raw_forecast = forecast_map.get(p.id)
            rows.append({
                'id':           p.id,
                'name':         p.display_name,
                'qty':          qty,
                'min_qty':      min_qty if has_min else None,
                'has_min':      has_min,
                'is_broken':    has_min and qty < (min_qty - 0.001),
                'qty_forecast': round(raw_forecast, 3) if raw_forecast is not None else None,
            })

        # KPIs sobre el conjunto completo
        kpis = {
            'total':  len(rows),
            'broken': sum(1 for r in rows if r['is_broken']),
            'ok':     sum(1 for r in rows if r['has_min'] and not r['is_broken']),
            'no_min': sum(1 for r in rows if not r['has_min']),
        }

        # Filtro
        if filter_type == 'broken':
            rows = [r for r in rows if r['is_broken']]
        elif filter_type == 'ok':
            rows = [r for r in rows if r['has_min'] and not r['is_broken']]
        elif filter_type == 'no_min':
            rows = [r for r in rows if not r['has_min']]

        # Sort
        _rev = (sort_dir == 'desc')
        if sort_field == 'name':
            rows.sort(key=lambda r: (r['name'] or '').lower(), reverse=_rev)
        elif sort_field == 'qty':
            rows.sort(key=lambda r: r['qty'], reverse=_rev)
        elif sort_field == 'min_qty':
            rows.sort(key=lambda r: (r['min_qty'] if r['min_qty'] is not None else -1), reverse=_rev)
        elif sort_field == 'qty_forecast':
            rows.sort(key=lambda r: r['qty_forecast'] if r['qty_forecast'] is not None else -999999, reverse=_rev)
        elif sort_field == 'status':
            rows.sort(key=lambda r: (0 if r['is_broken'] else 1 if not r['has_min'] else 2), reverse=_rev)
        else:
            # Default: quiebres primero, luego OK, luego sin mínimo; dentro de cada grupo por nombre
            rows.sort(key=lambda r: (
                0 if r['is_broken'] else 1 if not r['has_min'] else 2,
                (r['name'] or '').lower(),
            ))

        total_filtered = len(rows)
        offset = (max(1, page) - 1) * page_size
        page_rows = rows[offset:offset + page_size]

        # Conteo de OFs activas para los productos de esta página
        page_pids = [r['id'] for r in page_rows]
        if page_pids:
            mo_groups = self.env['mrp.production'].read_group(
                [('product_id', 'in', page_pids),
                 ('state', 'in', ['confirmed', 'progress', 'to_close'])] + no_subcontract_domain(self.env),
                ['product_id'],
                ['product_id'],
            )
            mo_count_map = {g['product_id'][0]: g['product_id_count'] for g in mo_groups}
        else:
            mo_count_map = {}
        for r in page_rows:
            r['mo_count'] = mo_count_map.get(r['id'], 0)

        # Tipos de producto para esta página (batch)
        if page_pids:
            page_pids_set = set(page_pids)
            page_prods = products.filtered(lambda p: p.id in page_pids_set)
            tmpl_type_map = {
                t.id: ', '.join(t.x_product_type_ids.mapped('name'))
                for t in page_prods.mapped('product_tmpl_id')
            }
            prod_to_tmpl = {p.id: p.product_tmpl_id.id for p in page_prods}
        else:
            tmpl_type_map = {}
            prod_to_tmpl = {}
        for r in page_rows:
            tmpl_id = prod_to_tmpl.get(r['id'])
            r['product_types'] = tmpl_type_map.get(tmpl_id, '') if tmpl_id else ''

        return {
            'error':          None,
            'kpis':           kpis,
            'products':       page_rows,
            'location_name':  location_name,
            'location_ids':   locations.ids,
            'location_id':    locations[0].id if locations else False,
            'total_filtered': total_filtered,
        }

    @api.model
    def get_product_mos_for_stock_break(self, product_id):
        """OFs activas de un producto para el acordeón del widget de quiebres de stock."""
        mos = self.env['mrp.production'].search([
            ('product_id', '=', product_id),
            ('state', 'in', ['confirmed', 'progress', 'to_close']),
        ] + no_subcontract_domain(self.env), limit=50, order='date_finished asc')
        state_labels = {
            'confirmed': 'Confirmada',
            'progress':  'En progreso',
            'to_close':  'Por cerrar',
        }
        return [{
            'id':            mo.id,
            'name':          mo.name,
            'state':         mo.state,
            'state_label':   state_labels.get(mo.state, mo.state),
            'product_qty':   round(mo.product_qty, 2),
            'qty_produced':  round(mo.qty_produced, 2),
            'uom':           mo.product_uom_id.name if mo.product_uom_id else '',
            'date_finished': mo.date_finished.strftime('%d/%m/%Y') if mo.date_finished else '—',
        } for mo in mos]

    @api.model
    def get_sales_chart_data(self, date_from, date_to, top_n=20, sale_category=None, product_categ_id=None, sort_by='qty', doc_type='sales'):
        tmpl_qty = {}
        tmpl_amount = {}

        if doc_type in ('sales', 'rfq', 'all'):
            so_states = []
            if doc_type in ('sales', 'all'):
                so_states += ['sale', 'done']
            if doc_type in ('rfq', 'all'):
                so_states += ['draft', 'sent']
            sol_domain = [
                ('order_id.state', 'in', so_states),
                ('order_id.date_order', '>=', date_from + ' 00:00:00'),
                ('order_id.date_order', '<=', date_to + ' 23:59:59'),
                ('product_id', '!=', False),
                ('product_id.sale_ok', '=', True),
            ]
            if product_categ_id:
                sol_domain.append(('product_id.categ_id', '=', int(product_categ_id)))
            try:
                groups = self.env['sale.order.line'].read_group(
                    sol_domain,
                    ['product_id', 'product_uom_qty:sum', 'price_subtotal:sum'],
                    ['product_id'],
                )
                for g in groups:
                    if not g['product_id']:
                        continue
                    pp = self.env['product.product'].browse(g['product_id'][0])
                    tid = pp.product_tmpl_id.id
                    tmpl_qty[tid] = tmpl_qty.get(tid, 0.0) + (g['product_uom_qty'] or 0.0)
                    tmpl_amount[tid] = tmpl_amount.get(tid, 0.0) + (g['price_subtotal'] or 0.0)
            except Exception:
                return []
        else:
            domain = [
                ('state', '=', 'done'),
                ('picking_id.picking_type_code', '=', 'outgoing'),
                ('date', '>=', date_from + ' 00:00:00'),
                ('date', '<=', date_to + ' 23:59:59'),
                ('product_id', '!=', False),
                ('product_id.sale_ok', '=', True),
            ]
            if product_categ_id:
                domain.append(('product_id.categ_id', '=', int(product_categ_id)))
            groups = self.env['stock.move.line'].read_group(
                domain,
                ['product_id', 'quantity:sum'],
                ['product_id'],
            )
            for g in groups:
                if not g['product_id']:
                    continue
                pp = self.env['product.product'].browse(g['product_id'][0])
                tid = pp.product_tmpl_id.id
                tmpl_qty[tid] = tmpl_qty.get(tid, 0.0) + (g['quantity'] or 0.0)

        if not tmpl_qty:
            return []

        if sale_category is not None and sale_category != '':
            tmpls_all = self.env['product.template'].browse(list(tmpl_qty.keys()))
            tmpl_by_id_f = {t.id: t for t in tmpls_all}
            if sale_category == '__none__':
                keep = {tid for tid in tmpl_qty
                        if not (tmpl_by_id_f.get(tid) and tmpl_by_id_f[tid].x_sale_category)}
            else:
                keep = {tid for tid in tmpl_qty
                        if tmpl_by_id_f.get(tid) and tmpl_by_id_f[tid].x_sale_category == sale_category}
            tmpl_qty   = {tid: q for tid, q in tmpl_qty.items()   if tid in keep}
            tmpl_amount = {tid: a for tid, a in tmpl_amount.items() if tid in keep}
            if not tmpl_qty:
                return []

        all_ids    = list(tmpl_qty.keys())
        templates  = self.env['product.template'].browse(all_ids)
        tmpl_by_id = {t.id: t for t in templates}

        rows = []
        for tid in all_ids:
            t = tmpl_by_id.get(tid)
            if not t:
                continue
            qty    = round(tmpl_qty[tid], 2)
            amount = round(tmpl_amount[tid], 2) if tid in tmpl_amount else round(qty * (t.list_price or 0.0), 2)
            rows.append({
                'tmpl_id':       tid,
                'name':          t.name,
                'code':          t.default_code or '',
                'sale_category': t.x_sale_category or '',
                'qty':           qty,
                'amount':        amount,
            })

        sort_key = 'amount' if sort_by == 'amount' else 'qty'
        result = sorted(rows, key=lambda r: r[sort_key], reverse=True)
        if top_n:
            result = result[:int(top_n)]
        return result

    @api.model
    def get_product_categories_for_chart(self):
        """Categorías de producto que tienen al menos un artículo vendible con entregas."""
        cats = self.env['product.category'].search([])
        result = []
        for c in cats:
            if self.env['product.template'].search_count([
                ('categ_id', '=', c.id),
                ('sale_ok', '=', True),
            ]):
                result.append({'id': c.id, 'name': c.complete_name})
        return sorted(result, key=lambda x: x['name'])

    @api.model
    def get_supplier_analysis_data(self, period_from, period_to, search=''):
        """Análisis de proveedores: cumplimiento, lead time, precio y volumen."""
        import calendar as _cal
        from datetime import date as _date

        def _parse_date(s, last_day=False):
            parts = s.split('-')
            y, m = int(parts[0]), int(parts[1])
            if len(parts) >= 3:
                return _date(y, m, int(parts[2]))
            return _date(y, m, _cal.monthrange(y, m)[1] if last_day else 1)

        try:
            d_from = _parse_date(period_from)
            d_to   = _parse_date(period_to, last_day=True)
        except Exception:
            return {'rows': [], 'kpis': {}, 'has_invoices': False}

        dt_from = fields.Datetime.to_string(datetime.combine(d_from, datetime.min.time()))
        dt_to   = fields.Datetime.to_string(datetime.combine(d_to,   datetime.max.time()))

        po_domain = [
            ('state', 'in', ['purchase', 'done']),
            ('date_approve', '>=', dt_from),
            ('date_approve', '<=', dt_to),
            ('company_id', '=', self.env.company.id),
        ]
        if search:
            po_domain.append(('partner_id.name', 'ilike', search))

        pos = self.env['purchase.order'].search(po_domain)
        _empty_kpis = {
            'supplier_count': 0, 'total_amount': 0, 'total_orders': 0,
            'avg_on_time_pct': None, 'avg_lead_time_days': None, 'avg_price_var_pct': None,
        }
        if not pos:
            return {'rows': [], 'kpis': _empty_kpis, 'has_invoices': False}

        # Aggregación por proveedor
        po_groups = self.env['purchase.order'].read_group(
            po_domain, ['partner_id', 'amount_total:sum'], ['partner_id'],
        )

        partner_data = {}
        for g in po_groups:
            if not g['partner_id']:
                continue
            pid = g['partner_id'][0]
            partner_data[pid] = {
                'partner_id':   pid,
                'partner_name': g['partner_id'][1],
                'order_count':  g['partner_id_count'],
                'total_amount': round(g['amount_total'] or 0.0, 2),
                'products':     set(),
                'pick_count':   0, 'on_time_count': 0,
                'delay_sum':    0.0, 'delay_count': 0,
                'complete_count': 0,
                'lt_sum':       0.0, 'lt_count': 0,
                'pvar_sum':     0.0, 'pvar_count': 0,
                'pending_inv':  0.0,
            }

        # Mapa po_id → (partner_id, date_approve)
        po_map = {po.id: (po.partner_id.id, po.date_approve) for po in pos}

        # Líneas de OC: productos + variación de precio
        po_line_data = self.env['purchase.order.line'].search_read(
            [('order_id', 'in', pos.ids), ('product_id', '!=', False)],
            ['order_id', 'product_id', 'price_unit'],
        )
        all_prod_ids = list({ln['product_id'][0] for ln in po_line_data if ln['product_id']})
        if all_prod_ids:
            std_map = {r['id']: r['standard_price']
                       for r in self.env['product.product'].search_read(
                           [('id', 'in', all_prod_ids)], ['id', 'standard_price'])}
        else:
            std_map = {}

        for ln in po_line_data:
            po_id    = ln['order_id'][0] if isinstance(ln['order_id'], (list, tuple)) else ln['order_id']
            prod_id  = ln['product_id'][0] if isinstance(ln['product_id'], (list, tuple)) else ln['product_id']
            partner_id = po_map.get(po_id, (None,))[0]
            if partner_id not in partner_data:
                continue
            pd = partner_data[partner_id]
            pd['products'].add(prod_id)
            std = std_map.get(prod_id, 0.0)
            if std > 0 and ln['price_unit'] > 0:
                pd['pvar_sum']   += (ln['price_unit'] - std) / std * 100
                pd['pvar_count'] += 1

        # Recepciones completadas vinculadas a las OCs
        pickings = self.env['stock.picking'].search([
            ('purchase_id', 'in', pos.ids),
            ('state', '=', 'done'),
            ('picking_type_code', '=', 'incoming'),
        ])
        # Recepciones que tuvieron backorder (parciales)
        partial_ids = set(
            self.env['stock.picking'].search(
                [('backorder_id', 'in', pickings.ids)]
            ).mapped('backorder_id').ids
        )

        for picking in pickings:
            po_id = picking.purchase_id.id if picking.purchase_id else None
            if not po_id or po_id not in po_map:
                continue
            partner_id, date_approve = po_map[po_id]
            if partner_id not in partner_data:
                continue
            pd = partner_data[partner_id]
            pd['pick_count'] += 1

            sched = picking.scheduled_date
            done  = picking.date_done
            if sched and done:
                delay = (done - sched).days
                if delay <= 0:
                    pd['on_time_count'] += 1
                else:
                    pd['delay_sum']   += delay
                    pd['delay_count'] += 1

            if picking.id not in partial_ids:
                pd['complete_count'] += 1

            if date_approve and done:
                lt = (done - date_approve).days
                if lt >= 0:
                    pd['lt_sum']   += lt
                    pd['lt_count'] += 1

        # Facturas pendientes (módulo account — opcional)
        has_invoices = False
        try:
            inv_groups = self.env['account.move'].read_group(
                [('move_type', '=', 'in_invoice'),
                 ('partner_id', 'in', list(partner_data.keys())),
                 ('payment_state', 'not in', ['paid', 'reversed']),
                 ('state', '=', 'posted'),
                 ('company_id', '=', self.env.company.id)],
                ['partner_id', 'amount_residual:sum'],
                ['partner_id'],
            )
            for g in inv_groups:
                if g['partner_id'] and g['partner_id'][0] in partner_data:
                    partner_data[g['partner_id'][0]]['pending_inv'] = round(g['amount_residual'] or 0.0, 2)
            has_invoices = True
        except Exception:
            pass

        # Construir filas
        rows = []
        for pid, d in partner_data.items():
            pc = d['pick_count']
            rows.append({
                'partner_id':        pid,
                'partner_name':      d['partner_name'],
                'order_count':       d['order_count'],
                'total_amount':      d['total_amount'],
                'distinct_products': len(d['products']),
                'pick_count':        pc,
                'on_time_pct':   round(d['on_time_count'] / pc * 100, 1) if pc > 0 else None,
                'avg_delay_days': round(d['delay_sum'] / d['delay_count'], 1) if d['delay_count'] > 0 else None,
                'complete_pct':  round(d['complete_count'] / pc * 100, 1) if pc > 0 else None,
                'avg_lead_time': round(d['lt_sum'] / d['lt_count'], 1) if d['lt_count'] > 0 else None,
                'avg_price_var_pct': round(d['pvar_sum'] / d['pvar_count'], 1) if d['pvar_count'] > 0 else None,
                'pending_inv':   d['pending_inv'] if has_invoices else None,
            })

        rows.sort(key=lambda r: r['total_amount'], reverse=True)

        def _wavg(rows, key):
            vals = [r[key] for r in rows if r[key] is not None]
            return round(sum(vals) / len(vals), 1) if vals else None

        total_pickings = sum(r['pick_count'] for r in rows)
        on_time_abs    = sum(d['on_time_count'] for d in partner_data.values())
        kpis = {
            'supplier_count':     len(rows),
            'total_amount':       round(sum(r['total_amount'] for r in rows), 2),
            'total_orders':       sum(r['order_count'] for r in rows),
            'avg_on_time_pct':    round(on_time_abs / total_pickings * 100, 1) if total_pickings > 0 else None,
            'avg_lead_time_days': _wavg(rows, 'avg_lead_time'),
            'avg_price_var_pct':  _wavg(rows, 'avg_price_var_pct'),
        }

        cfg = self.env['mrp.reschedule.config'].search([], limit=1)
        show_supplier_cat = bool(cfg and cfg.enable_supplier_categories)
        if show_supplier_cat:
            cat_map = {r['id']: r['x_supplier_category'] or ''
                       for r in self.env['res.partner'].search_read(
                           [('id', 'in', list(partner_data.keys()))],
                           ['id', 'x_supplier_category']
                       )}
            for row in rows:
                row['supplier_cat'] = cat_map.get(row['partner_id'], '')

        sup_config = {
            'sup_on_time_green':   cfg.sup_on_time_green_pct   if cfg else 90,
            'sup_on_time_yellow':  cfg.sup_on_time_yellow_pct  if cfg else 70,
            'sup_delay_green':     cfg.sup_delay_green_days    if cfg else 1,
            'sup_delay_yellow':    cfg.sup_delay_yellow_days   if cfg else 3,
            'sup_complete_green':  cfg.sup_complete_green_pct  if cfg else 95,
            'sup_complete_yellow': cfg.sup_complete_yellow_pct if cfg else 80,
            'sup_price_var_green':  cfg.sup_price_var_green_pct  if cfg else 3.0,
            'sup_price_var_yellow': cfg.sup_price_var_yellow_pct if cfg else 10.0,
        }

        return {
            'rows': rows, 'kpis': kpis, 'has_invoices': has_invoices,
            'config': sup_config, 'show_supplier_cat': show_supplier_cat,
        }

    @api.model
    def get_supplier_pos_for_analysis(self, partner_id, period_from, period_to):
        """OCs de un proveedor para el acordeón del widget de análisis de proveedores."""
        import calendar as _cal
        from datetime import date as _date
        def _parse_date(s, last_day=False):
            parts = s.split('-')
            y, m = int(parts[0]), int(parts[1])
            if len(parts) >= 3:
                return _date(y, m, int(parts[2]))
            return _date(y, m, _cal.monthrange(y, m)[1] if last_day else 1)

        try:
            d_from = _parse_date(period_from)
            d_to   = _parse_date(period_to, last_day=True)
        except Exception:
            return []

        dt_from = fields.Datetime.to_string(datetime.combine(d_from, datetime.min.time()))
        dt_to   = fields.Datetime.to_string(datetime.combine(d_to,   datetime.max.time()))

        pos = self.env['purchase.order'].search([
            ('partner_id', '=', partner_id),
            ('state', 'in', ['purchase', 'done']),
            ('date_order', '>=', dt_from),
            ('date_order', '<=', dt_to),
            ('company_id', '=', self.env.company.id),
        ], order='date_order desc')

        rows = []
        for po in pos:
            pickings = po.picking_ids.filtered(
                lambda p: p.state != 'cancel' and p.picking_type_code == 'incoming'
            )
            done_picks = pickings.filtered(lambda p: p.state == 'done')
            if not pickings:
                receipt_status = 'none'
            elif len(done_picks) == len(pickings):
                receipt_status = 'full'
            elif done_picks:
                receipt_status = 'partial'
            else:
                receipt_status = 'pending'

            rows.append({
                'po_id':          po.id,
                'name':           po.name,
                'date_approve':   po.date_approve.strftime('%d/%m/%Y') if po.date_approve else '',
                'date_planned':   po.date_planned.strftime('%d/%m/%Y') if po.date_planned else '',
                'amount_total':   round(po.amount_total, 2),
                'product_count':  len(po.order_line.mapped('product_id')),
                'receipt_status': receipt_status,
            })
        return rows
