# Copyright (C) 2024 - MRP Planner
# License LGPL-3.0 or later (https://www.gnu.org/licenses/lgpl).
"""
Módulo: mrp_planner_dashboard_forecast.py
Modelo: extensión de 'mrp.planner.dashboard'

Extiende el dashboard del planificador MRP con toda la lógica de Forecast:
consulta líneas de forecast, ÓFs planificadas, entregas y demanda real de ventas
para construir una tabla pivotada por producto × mes con KPIs de cobertura,
precisión de forecast y rotación de stock.

Responsabilidades:
- Exponer la lista de almacenes disponibles para el filtro de forecast.
- Calcular y retornar el payload completo del widget de forecast (KPIs + tabla).
- Proveer el drill-down de ÓFs por producto para el acordeón del widget.
- Generar y devolver un archivo Excel (.xlsx) con el resumen de forecast vs. OFs.

Relacionado con:
- mrp.planner.dashboard: clase base que este mixin extiende con _inherit.
- mrp.forecast.line: líneas de forecast agrupadas por producto y período.
- mrp.production: órdenes de fabricación consultadas para cobertura de forecast.
- mrp.reschedule.config: configuración de umbrales, fórmula de precisión y estados de OF.
- stock.move.line: movimientos de salida completados (entregas reales).
- sale.order.line: demanda real de pedidos de venta confirmados.
- stock.quant: stock actual para cálculo de rotación.
"""
import logging
import pytz
import io
import base64
from datetime import datetime, timedelta

from odoo import models, fields, api
from odoo.addons.odoo_mrp_planner.models.mrp_schedule_mixin import no_subcontract_domain

_logger = logging.getLogger(__name__)


class MrpPlannerDashboardForecast(models.TransientModel):
    _inherit = 'mrp.planner.dashboard'

    # ── Forecast ─────────────────────────────────────────────────────────────

    @api.model
    def get_warehouses_for_forecast(self):
        """
        Retorna la lista de almacenes accesibles por el usuario actual para el filtro de forecast.

        Si el usuario tiene el flag 'mrp_planner_all_warehouses' activo, o no tiene
        almacenes asignados, se devuelven todos los almacenes de la base de datos.
        En caso contrario, solo se devuelven los almacenes explícitamente asignados al usuario.

        :returns: list[dict] — lista de dicts con 'id' (int) y 'name' (str) de cada almacén,
                  ordenados alfabéticamente por nombre.
        """
        user = self.env.user
        if user.mrp_planner_all_warehouses or not user.mrp_planner_warehouse_ids:
            whs = self.env['stock.warehouse'].search([], order='name')
        else:
            whs = user.mrp_planner_warehouse_ids.sorted('name')
        return [{'id': w.id, 'name': w.name} for w in whs]

    @api.model
    def get_forecast_dashboard_data(self, period_from, period_to, warehouse_ids=None):
        """
        Devuelve KPIs y tabla pivotada forecast vs ÓFs para el rango de meses indicado.

        Construye el payload completo que consume el widget de forecast del dashboard.
        El proceso interno sigue este orden:
        1. Parsear rango de fechas y convertir a UTC para dominios Datetime.
        2. Leer configuración (umbrales, fórmula de precisión, estados de OF aceptados).
        3. Consultar mrp.forecast.line en el período.
        4. Consultar mrp.production filtradas por estado, fecha y exclusión de subcontratación.
        5. Consultar stock.move.line (entregas completadas) para los productos con forecast.
        6. Consultar sale.order.line (demanda real) para los productos con forecast.
        7. Consultar stock.quant (snapshot de stock actual) por almacén si se filtra.
        8. Construir filas por producto con celdas mensuales y totales.
        9. Calcular KPIs globales y métricas para productos sin línea de forecast.

        :param period_from: str — fecha de inicio en formato 'YYYY-MM' o 'YYYY-MM-DD'.
        :param period_to:   str — fecha de fin en formato 'YYYY-MM' o 'YYYY-MM-DD'.
        :param warehouse_ids: list[int] | None — IDs de almacenes para filtrar stock y OFs.
                              Si es None o lista vacía, se consideran todos los almacenes.
        :returns: dict con las claves:
            - 'kpis' (dict): métricas globales del período.
            - 'months' (list[str]): lista de meses 'YYYY-MM' en el rango.
            - 'month_totals' (list[dict]): totales por mes de forecast, OFs, entregado y demanda.
            - 'rows' (list[dict]): una fila por producto con 'cells', totales y métricas de precisión.
            - 'warning_pct' (float): umbral de alerta configurado (defecto 70 %).
            - 'critical_pct' (float): umbral crítico configurado (defecto 50 %).
            - 'rotation_unit' (str): unidad de rotación 'days' o 'months'.
            - 'acc_formula' (str): fórmula de precisión activa ('simple', 'mape', 'wape', 'wmape', 'bias').
            - 'mo_states' (list[str]): estados de OF incluidos en el cálculo.
        """
        from datetime import date as _date
        import calendar as _calendar

        warehouse_ids = warehouse_ids or []

        def _parse_ym(ym):
            """Convierte 'YYYY-MM' o 'YYYY-MM-DD' a un objeto date (día 1 si no se especifica)."""
            parts = ym.split('-')
            y, m = int(parts[0]), int(parts[1])
            d = int(parts[2]) if len(parts) >= 3 else 1
            return _date(y, m, d)

        def _months_between(d_from, d_to):
            """Genera lista de strings 'YYYY-MM' para cada mes entre d_from y d_to inclusive."""
            months = []
            d = _date(d_from.year, d_from.month, 1)
            while d <= _date(d_to.year, d_to.month, 1):
                months.append(f"{d.year}-{d.month:02d}")
                if d.month == 12:
                    d = _date(d.year + 1, 1, 1)
                else:
                    d = _date(d.year, d.month + 1, 1)
            return months

        try:
            d_from = _parse_ym(period_from)
            d_to   = _parse_ym(period_to)
        except Exception:
            return {'kpis': {}, 'months': [], 'month_totals': [], 'rows': [],
                    'warning_pct': 70, 'critical_pct': 50}

        months = _months_between(d_from, d_to)

        cfg = self.env['mrp.reschedule.config'].search([], limit=1)
        warning_pct    = cfg.forecast_warning_pct    if cfg else 70   # 70 %: umbral de alerta por defecto (cobertura aceptable mínima)
        critical_pct   = cfg.forecast_critical_pct   if cfg else 50   # 50 %: umbral crítico por defecto (cobertura insuficiente)
        rotation_unit   = (cfg.forecast_rotation_unit   if cfg else None) or 'days'
        rotation_method = (cfg.forecast_rotation_method if cfg else None) or 'units'
        acc_formula      = (cfg.forecast_acc_formula      if cfg else None) or 'simple'
        precision_source = (cfg.forecast_precision_source if cfg else None) or 'demand'
        coverage_unit            = (cfg.forecast_coverage_unit            if cfg else None) or 'days'
        coverage_demand_source   = (cfg.forecast_coverage_demand_source   if cfg else None) or 'forecast'
        coverage_alerts_enabled  = bool(cfg.forecast_coverage_alerts_enabled) if cfg else False
        coverage_warn_days       = (cfg.forecast_coverage_warn_days       if cfg else None) or 30
        coverage_critical_days   = (cfg.forecast_coverage_critical_days   if cfg else None) or 15
        mo_coverage_show_pct     = bool(cfg.forecast_mo_coverage_show_pct) if cfg else True
        mo_coverage_denominator  = (cfg.forecast_mo_coverage_denominator  if cfg else None) or 'forecast'
        mo_coverage_color_scope  = (cfg.forecast_mo_coverage_color_scope  if cfg else None) or 'both'

        # Estados de OF configurados
        mo_states = []
        if cfg:
            if cfg.forecast_mo_state_draft:     mo_states.append('draft')
            if cfg.forecast_mo_state_confirmed: mo_states.append('confirmed')
            if cfg.forecast_mo_state_progress:  mo_states.append('progress')
            if cfg.forecast_mo_state_to_close:  mo_states.append('to_close')
            if cfg.forecast_mo_state_done:      mo_states.append('done')
        if not mo_states:
            # Fallback a los estados productivos más relevantes si la config no especifica ninguno
            mo_states = ['confirmed', 'progress', 'to_close']

        last_day_of_to = d_to  # día exacto seleccionado por el usuario

        # Conversión a UTC para dominios Datetime (Odoo guarda en UTC)
        tz_name  = self.env.context.get('tz') or self.env.user.tz or 'UTC'
        user_tz  = pytz.timezone(tz_name)

        def _to_utc(dt_naive):
            """Localiza un datetime naive en la zona del usuario y lo convierte a UTC naive."""
            return user_tz.localize(dt_naive).astimezone(pytz.utc).replace(tzinfo=None)

        def _dt_ym(dt_utc):
            """Devuelve 'YYYY-MM' en hora local a partir de un Datetime UTC."""
            return pytz.utc.localize(dt_utc).astimezone(user_tz).strftime('%Y-%m')

        dt_from = _to_utc(datetime.combine(d_from, datetime.min.time()))
        dt_to   = _to_utc(datetime.combine(last_day_of_to, datetime.max.time()))

        # ── Forecast lines ────────────────────────────────────────────────────
        fc_domain = [
            ('period', '>=', _date(d_from.year, d_from.month, 1)),
            ('period', '<=', last_day_of_to),
            ('company_id', '=', self.env.company.id),
        ]

        fc_lines = self.env['mrp.forecast.line'].search(fc_domain)

        # Precarga en un solo SELECT todos los campos relacionales necesarios
        _product_ids_fc = fc_lines.mapped('product_id')
        _product_read   = {r['id']: r for r in _product_ids_fc.read(['id', 'display_name', 'product_tmpl_id'])}

        # Estructura: {product_id: {month_str: forecast_qty}}
        fc_data = {}
        for line in fc_lines:
            pid = line.product_id.id
            ym  = f"{line.period.year}-{line.period.month:02d}"
            if pid not in fc_data:
                _pr = _product_read.get(pid, {})
                fc_data[pid] = {
                    'product':         _pr.get('display_name', ''),
                    'product_tmpl_id': _pr.get('product_tmpl_id', [False])[0],
                }
            fc_data[pid][ym] = fc_data[pid].get(ym, 0.0) + line.forecast_qty

        # ── ÓFs planificadas ──────────────────────────────────────────────────
        mo_mode = (cfg.comparison_date_mode if cfg else None) or 'finish_date'

        no_sc_domain = no_subcontract_domain(self.env)
        wh_filter = [('picking_type_id.warehouse_id', 'in', warehouse_ids)] if warehouse_ids else []

        mo_data = {}  # {product_id: {month_str: qty}}

        if mo_mode == 'finish_date':
            mo_domain = [
                ('state', 'in', mo_states),
                ('date_finished', '>=', fields.Datetime.to_string(dt_from)),
                ('date_finished', '<=', fields.Datetime.to_string(dt_to)),
            ] + no_sc_domain + wh_filter
            mos = self.env['mrp.production'].search(mo_domain)
            for _mo in mos.read(['product_id', 'date_finished', 'product_qty']):
                pid = _mo['product_id'][0] if _mo['product_id'] else False
                df  = _mo['date_finished']
                if not pid or not df:
                    continue
                ym = _dt_ym(df)
                if ym not in months:
                    continue
                if pid not in mo_data:
                    mo_data[pid] = {}
                mo_data[pid][ym] = mo_data[pid].get(ym, 0.0) + _mo['product_qty']
        else:
            # overlap y proportional: OFs que solapan el rango completo
            mo_domain_wide = [
                ('state', 'in', mo_states),
                ('date_start', '<=', fields.Datetime.to_string(dt_to)),
                '|',
                ('date_finished', '>=', fields.Datetime.to_string(dt_from)),
                ('date_finished', '=', False),
            ] + no_sc_domain + wh_filter
            mos_wide = self.env['mrp.production'].search(mo_domain_wide)

            # Límites UTC de cada mes para cálculo de solapamiento
            def _month_bounds_utc(ym_str):
                y, m = int(ym_str[:4]), int(ym_str[5:])
                m_start = _to_utc(datetime(y, m, 1, 0, 0, 0))
                if m == 12:
                    m_end = _to_utc(datetime(y + 1, 1, 1, 0, 0, 0)) - timedelta(seconds=1)
                else:
                    m_end = _to_utc(datetime(y, m + 1, 1, 0, 0, 0)) - timedelta(seconds=1)
                return m_start, m_end

            month_bounds = {ym: _month_bounds_utc(ym) for ym in months}

            for _mo in mos_wide.read(['product_id', 'date_start', 'date_finished', 'product_qty']):
                pid = _mo['product_id'][0] if _mo['product_id'] else False
                if not pid:
                    continue
                mo_start = _mo['date_start']
                mo_end   = _mo['date_finished']
                qty      = _mo['product_qty']

                if mo_mode == 'overlap':
                    for ym, (m_start, m_end) in month_bounds.items():
                        if mo_start and mo_start <= m_end and (not mo_end or mo_end >= m_start):
                            if pid not in mo_data:
                                mo_data[pid] = {}
                            mo_data[pid][ym] = mo_data[pid].get(ym, 0.0) + qty
                else:  # proportional
                    if mo_start and mo_end and mo_start < mo_end:
                        total_secs = (mo_end - mo_start).total_seconds()
                        for ym, (m_start, m_end) in month_bounds.items():
                            ov_start = max(mo_start, m_start)
                            ov_end   = min(mo_end, m_end)
                            overlap_secs = max(0.0, (ov_end - ov_start).total_seconds())
                            if overlap_secs > 0:
                                if pid not in mo_data:
                                    mo_data[pid] = {}
                                mo_data[pid][ym] = mo_data[pid].get(ym, 0.0) + qty * (overlap_secs / total_secs)
                    elif mo_end:
                        # sin date_start: fallback a mes de cierre
                        ym = _dt_ym(mo_end)
                        if ym in months:
                            if pid not in mo_data:
                                mo_data[pid] = {}
                            mo_data[pid][ym] = mo_data[pid].get(ym, 0.0) + qty

        def _cov_days(stock, period_days, demand):
            return round(stock * period_days / demand, 1) if demand > 0 else None

        def _cov_months(stock, period_days, demand):
            return round(stock * period_days / demand / 30, 1) if demand > 0 else None

        # ── Ids de productos con forecast ──────────────────────────────────────
        all_product_ids      = set(fc_data.keys())
        all_product_ids_list = list(all_product_ids)
        # n_months para cálculo de rotación: duración real en meses, no meses de calendario tocados.
        # Ej: 08/04 → 08/07 = 91 días ≈ 3,03 meses, pero len(months) = 4 (abr, may, jun, jul).
        _period_days = max(1, (last_day_of_to - d_from).days)
        n_months     = max(1.0, _period_days / 30.0)

        # ── Movimientos de salida completados (entregado) ──────────────────────
        del_line_domain = [
            ('state', '=', 'done'),
            ('picking_id.picking_type_id.code', '=', 'outgoing'),
            ('date', '>=', fields.Datetime.to_string(dt_from)),
            ('date', '<=', fields.Datetime.to_string(dt_to)),
            ('product_id', 'in', all_product_ids_list),
            ('company_id', '=', self.env.company.id),
        ]
        del_data = {}   # {product_id: {ym: qty}}
        for _ml in self.env['stock.move.line'].search(del_line_domain).read(
                ['product_id', 'date', 'quantity']):
            pid = _ml['product_id'][0] if _ml['product_id'] else False
            dt  = _ml['date']
            if not pid or not dt:
                continue
            ym = _dt_ym(dt)
            if ym not in months:
                continue
            del_data.setdefault(pid, {})
            del_data[pid][ym] = del_data[pid].get(ym, 0.0) + _ml['quantity']

        # ── Demanda real: pedidos de venta confirmados ─────────────────────────
        so_data = {}    # {product_id: {ym: qty}}
        try:
            so_domain = [
                ('order_id.state', 'in', ('sale', 'done')),
                ('order_id.date_order', '>=', fields.Datetime.to_string(dt_from)),
                ('order_id.date_order', '<=', fields.Datetime.to_string(dt_to)),
                ('product_id', 'in', all_product_ids_list),
                ('company_id', '=', self.env.company.id),
            ]
            sol_rows = self.env['sale.order.line'].search(so_domain).read(
                ['product_id', 'product_uom_qty', 'order_id']
            )
            # Precarga date_order de todas las sale.order en un solo SELECT
            _order_ids  = list({r['order_id'][0] for r in sol_rows if r['order_id']})
            _order_dates = {o['id']: o['date_order'] for o in
                            self.env['sale.order'].browse(_order_ids).read(['id', 'date_order'])}
            for _sl in sol_rows:
                pid = _sl['product_id'][0] if _sl['product_id'] else False
                if not pid:
                    continue
                dt  = _order_dates.get(_sl['order_id'][0]) if _sl['order_id'] else None
                if not dt:
                    continue
                ym  = _dt_ym(dt)
                if ym not in months:
                    continue
                so_data.setdefault(pid, {})
                so_data[pid][ym] = so_data[pid].get(ym, 0.0) + _sl['product_uom_qty']
        except Exception:
            pass    # módulo sale no disponible

        # ── Stock actual (snapshot) ───────────────────────────────────────────
        stock_data = {}   # {product_id: qty}
        quant_domain = [
            ('location_id.usage', '=', 'internal'),
            ('product_id', 'in', all_product_ids_list),
            ('company_id', '=', self.env.company.id),
        ]
        if warehouse_ids:
            wh_recs  = self.env['stock.warehouse'].browse(warehouse_ids)
            loc_ids  = wh_recs.mapped('lot_stock_id').ids
            if loc_ids:
                quant_domain.append(('location_id', 'in', loc_ids))
        for _qg in self.env['stock.quant'].read_group(
                quant_domain, ['product_id', 'quantity:sum'], ['product_id']):
            pid = _qg['product_id'][0] if _qg['product_id'] else False
            if pid:
                stock_data[pid] = round(_qg['quantity'] or 0.0, 6)

        # ── Datos de rotación (batch, según método) ───────────────────────────────
        period_days_rot = n_months * 30  # días comerciales del período (aproximación de mes comercial)
        dt_from_str = fields.Datetime.to_string(dt_from)
        dt_to_str   = fields.Datetime.to_string(dt_to)

        # Método units: reconstrucción de stock histórico desde stock.move (sin valorización)
        qty_in_start_by_pid  = {}
        qty_out_start_by_pid = {}
        qty_in_end_by_pid    = {}
        qty_out_end_by_pid   = {}

        if rotation_method == 'units' and all_product_ids_list:
            try:
                # sudo(): usuario no tiene acceso directo a stock.move; se lee sólo el agregado para el dashboard
                SM = self.env['stock.move'].sudo()
                _sm_base = [
                    ('state', '=', 'done'),
                    ('product_id', 'in', all_product_ids_list),
                    ('company_id', '=', self.env.company.id),
                ]
                for g in SM.read_group(_sm_base + [
                    ('date', '<', dt_from_str),
                    ('location_dest_id.usage', '=', 'internal'),
                    ('location_id.usage', '!=', 'internal'),
                ], ['product_id', 'product_qty:sum'], ['product_id']):
                    if g['product_id']:
                        qty_in_start_by_pid[g['product_id'][0]] = g['product_qty'] or 0.0

                for g in SM.read_group(_sm_base + [
                    ('date', '<', dt_from_str),
                    ('location_id.usage', '=', 'internal'),
                    ('location_dest_id.usage', '!=', 'internal'),
                ], ['product_id', 'product_qty:sum'], ['product_id']):
                    if g['product_id']:
                        qty_out_start_by_pid[g['product_id'][0]] = g['product_qty'] or 0.0

                for g in SM.read_group(_sm_base + [
                    ('date', '<=', dt_to_str),
                    ('location_dest_id.usage', '=', 'internal'),
                    ('location_id.usage', '!=', 'internal'),
                ], ['product_id', 'product_qty:sum'], ['product_id']):
                    if g['product_id']:
                        qty_in_end_by_pid[g['product_id'][0]] = g['product_qty'] or 0.0

                for g in SM.read_group(_sm_base + [
                    ('date', '<=', dt_to_str),
                    ('location_id.usage', '=', 'internal'),
                    ('location_dest_id.usage', '!=', 'internal'),
                ], ['product_id', 'product_qty:sum'], ['product_id']):
                    if g['product_id']:
                        qty_out_end_by_pid[g['product_id'][0]] = g['product_qty'] or 0.0
            except Exception:
                pass

        cogs_by_pid      = {}
        inv_start_by_pid = {}
        inv_end_by_pid   = {}
        sales_rev_by_pid = {}

        if rotation_method in ('cogs', 'sales') and 'stock.valuation.layer' in self.env:
            try:
                # sudo(): usuario no tiene acceso directo a stock.valuation.layer; se lee sólo el agregado para el dashboard
                SVL = self.env['stock.valuation.layer'].sudo()
                # COGS: capas con valor negativo (salidas) dentro del período
                for g in SVL.read_group([
                    ('product_id', 'in', all_product_ids_list),
                    ('create_date', '>=', dt_from_str),
                    ('create_date', '<=', dt_to_str),
                    ('value', '<', 0),
                    ('company_id', '=', self.env.company.id),
                ], ['product_id', 'value:sum'], ['product_id']):
                    if g['product_id']:
                        cogs_by_pid[g['product_id'][0]] = -(g['value'] or 0.0)

                # Inventario valorizado acumulado al inicio del período
                for g in SVL.read_group([
                    ('product_id', 'in', all_product_ids_list),
                    ('create_date', '<', dt_from_str),
                    ('company_id', '=', self.env.company.id),
                ], ['product_id', 'value:sum'], ['product_id']):
                    if g['product_id']:
                        inv_start_by_pid[g['product_id'][0]] = g['value'] or 0.0

                # Inventario valorizado acumulado al fin del período
                for g in SVL.read_group([
                    ('product_id', 'in', all_product_ids_list),
                    ('create_date', '<=', dt_to_str),
                    ('company_id', '=', self.env.company.id),
                ], ['product_id', 'value:sum'], ['product_id']):
                    if g['product_id']:
                        inv_end_by_pid[g['product_id'][0]] = g['value'] or 0.0
            except Exception:
                pass

        if rotation_method == 'sales':
            try:
                # sudo(): usuario no tiene acceso directo a sale.order.line; se lee sólo el agregado para el dashboard
                for g in self.env['sale.order.line'].sudo().read_group([
                    ('order_id.state', 'in', ('sale', 'done')),
                    ('order_id.date_order', '>=', fields.Datetime.to_string(dt_from)),
                    ('order_id.date_order', '<=', fields.Datetime.to_string(dt_to)),
                    ('product_id', 'in', all_product_ids_list),
                    ('company_id', '=', self.env.company.id),
                ], ['product_id', 'price_subtotal:sum'], ['product_id']):
                    if g['product_id']:
                        sales_rev_by_pid[g['product_id'][0]] = g['price_subtotal'] or 0.0
            except Exception:
                pass

        # ── Construir filas ────────────────────────────────────────────────────
        rows = []
        # Categorías de venta por product.template
        tmpl_ids = [fc_data[pid].get('product_tmpl_id') for pid in all_product_ids
                    if fc_data[pid].get('product_tmpl_id')]
        if tmpl_ids:
            tmpl_info = {}
            _tmpl_rows = self.env['product.template'].browse(tmpl_ids).read(
                ['id', 'x_sale_category', 'categ_id', 'x_product_type_ids']
            )
            # Precarga nombres de categorías en un solo SELECT
            _categ_ids = list({r['categ_id'][0] for r in _tmpl_rows if r['categ_id']})
            _categ_names = {c['id']: c['name'] for c in
                            self.env['product.category'].browse(_categ_ids).read(['id', 'name'])}
            # Precarga nombres de tipos de producto
            _type_ids_all = list({tid for r in _tmpl_rows for tid in (r['x_product_type_ids'] or [])})
            _type_names = {tp['id']: tp['name'] for tp in
                           self.env['x.product.type'].browse(_type_ids_all).read(['id', 'name'])} \
                          if _type_ids_all else {}
            for _tr in _tmpl_rows:
                _categ_id = _tr['categ_id'][0] if _tr['categ_id'] else False
                tmpl_info[_tr['id']] = {
                    'sale_category': _tr.get('x_sale_category') or '',
                    'product_categ': _categ_names.get(_categ_id, '') if _categ_id else '',
                    'product_types': ', '.join(_type_names.get(tid, '') for tid in (_tr['x_product_type_ids'] or [])),
                }
        else:
            tmpl_info = {}

        for pid in all_product_ids:
            pname    = fc_data[pid]['product']
            pid_del  = del_data.get(pid, {})
            pid_so   = so_data.get(pid, {})
            stock_qty = round(stock_data.get(pid, 0.0), 2)
            cells            = []
            tot_fc           = 0.0
            tot_mos          = 0.0
            tot_del          = 0.0
            tot_so           = 0.0
            _mape_acc_sum    = 0.0   # Σ precisión por período (MAPE)
            _mape_acc_count  = 0     # períodos con del > 0 (MAPE)
            _wape_abs_err    = 0.0   # Σ|error| ponderado por real (WAPE)
            _wmape_abs_err   = 0.0   # Σ|error| ponderado por forecast (WMAPE)

            for ym in months:
                fc_qty  = fc_data[pid].get(ym, 0.0)
                mo_qty  = mo_data.get(pid, {}).get(ym, 0.0)
                del_qty = pid_del.get(ym, 0.0)
                so_qty  = pid_so.get(ym, 0.0)
                pct      = round(mo_qty  / fc_qty * 100, 1) if fc_qty > 0 else 0.0
                svc_rate = round(del_qty / so_qty * 100, 1) if so_qty > 0 else None

                actual  = del_qty if precision_source == 'delivery' else so_qty
                abs_err = abs(actual - fc_qty)
                if actual > 0:
                    _mape_acc_sum   += max(0.0, 100.0 - abs_err / actual * 100)
                    _mape_acc_count += 1
                    _wape_abs_err   += abs_err
                if fc_qty > 0:
                    _wmape_abs_err  += abs_err
                # Valor de celda solo para la fórmula configurada
                if acc_formula in ('mape', 'wape'):
                    fc_acc = round(max(0.0, 100.0 - abs_err / actual * 100), 1) if actual > 0 else None
                elif acc_formula == 'wmape':
                    fc_acc = round(max(0.0, 100.0 - abs_err / fc_qty * 100), 1) if fc_qty > 0 else None
                elif acc_formula == 'bias':
                    fc_acc = round((actual - fc_qty) / fc_qty * 100, 1) if fc_qty > 0 else None
                else:  # simple
                    fc_acc = round(actual / fc_qty * 100, 1) if fc_qty > 0 else None

                demand_gap_pct = round((so_qty - fc_qty) / fc_qty * 100, 1) if fc_qty > 0 else None

                cells.append({
                    'month':           ym,
                    'forecast':        round(fc_qty,  2),
                    'mos':             round(mo_qty,  2),
                    'pct':             pct,
                    'delivered':       round(del_qty, 2),
                    'so_demand':       round(so_qty,  2),
                    'service_rate':    svc_rate,
                    'forecast_acc':    fc_acc,
                    'demand_gap_pct':  demand_gap_pct,
                })
                tot_fc  += fc_qty
                tot_mos += mo_qty
                tot_del += del_qty
                tot_so  += so_qty

            tot_pct = round(tot_mos / tot_fc * 100, 1) if tot_fc > 0 else 0.0
            tot_svc = round(tot_del / tot_so * 100, 1) if tot_so > 0 else None
            if acc_formula == 'mape':
                tot_acc = round(_mape_acc_sum / _mape_acc_count, 1) if _mape_acc_count > 0 else None
            elif acc_formula == 'wape':
                tot_acc = round(max(0.0, 100.0 - _wape_abs_err / tot_so * 100), 1) if tot_so > 0 else None
            elif acc_formula == 'wmape':
                tot_acc = round(max(0.0, 100.0 - _wmape_abs_err / tot_fc * 100), 1) if tot_fc > 0 else None
            elif acc_formula == 'bias':
                tot_acc = round((tot_so - tot_fc) / tot_fc * 100, 1) if tot_fc > 0 else None
            else:
                tot_acc = round(tot_so / tot_fc * 100, 1) if tot_fc > 0 else None

            rot_months = None
            rot_days   = None
            avg_stock_qty = stock_qty  # fallback: stock actual si no hay historial
            if rotation_method == 'units':
                stock_start = max(0.0, qty_in_start_by_pid.get(pid, 0.0) - qty_out_start_by_pid.get(pid, 0.0))
                stock_end   = max(0.0, qty_in_end_by_pid.get(pid, 0.0)   - qty_out_end_by_pid.get(pid, 0.0))
                if stock_start > 0 or stock_end > 0:
                    avg_stock_qty = (stock_start + stock_end) / 2.0
                avg_monthly_del = tot_del / n_months
                if avg_monthly_del > 0:
                    rot_months = round(avg_stock_qty / avg_monthly_del, 1)
                    rot_days   = int(round(avg_stock_qty / avg_monthly_del * 30))
            elif rotation_method in ('cogs', 'sales'):
                inv_s   = inv_start_by_pid.get(pid, 0.0)
                inv_e   = inv_end_by_pid.get(pid, 0.0)
                avg_inv = (inv_s + inv_e) / 2.0
                if avg_inv > 0:
                    base = cogs_by_pid.get(pid, 0.0) if rotation_method == 'cogs' \
                           else sales_rev_by_pid.get(pid, 0.0)
                    if base > 0:
                        dio        = period_days_rot * avg_inv / base
                        rot_days   = int(round(dio))
                        rot_months = round(dio / 30.0, 1)

            rows.append({
                'product_id':         pid,
                'product_tmpl_id':    fc_data[pid].get('product_tmpl_id'),
                'product':            pname,
                'cells':              cells,
                'stock_qty':          stock_qty,
                'avg_stock_qty':      round(avg_stock_qty, 2),
                'rotation_days':      rot_days,
                'rotation_months':    rot_months,
                'coverage_days':      _cov_days(stock_qty, _period_days,
                                          {'forecast': tot_fc, 'so_demand': tot_so, 'delivered': tot_del}.get(coverage_demand_source, tot_fc)),
                'coverage_months':    _cov_months(stock_qty, _period_days,
                                          {'forecast': tot_fc, 'so_demand': tot_so, 'delivered': tot_del}.get(coverage_demand_source, tot_fc)),
                'total_forecast':     round(tot_fc,  2),
                'total_mos':          round(tot_mos, 2),
                'total_pct':          tot_pct,
                'total_delivered':    round(tot_del, 2),
                'total_so_demand':    round(tot_so,  2),
                'total_service_rate': tot_svc,
                'total_forecast_acc': tot_acc,
                'demand_gap_pct': round((tot_so - tot_fc) / tot_fc * 100, 1) if tot_fc > 0 else None,
                'acc_all': {
                    'simple': round(tot_so / tot_fc * 100, 1) if tot_fc > 0 else None,
                    'mape':   round(_mape_acc_sum / _mape_acc_count, 1) if _mape_acc_count > 0 else None,
                    'wape':   round(max(0.0, 100.0 - _wape_abs_err / tot_so * 100), 1) if tot_so > 0 else None,
                    'wmape':  round(max(0.0, 100.0 - _wmape_abs_err / tot_fc * 100), 1) if tot_fc > 0 else None,
                    'bias':   round((tot_so - tot_fc) / tot_fc * 100, 1) if tot_fc > 0 else None,
                },
                'sale_category':      tmpl_info.get(fc_data[pid].get('product_tmpl_id'), {}).get('sale_category', ''),
                'product_categ':      tmpl_info.get(fc_data[pid].get('product_tmpl_id'), {}).get('product_categ', ''),
                'product_types':      tmpl_info.get(fc_data[pid].get('product_tmpl_id'), {}).get('product_types', ''),
                '_mape_acc_sum':      _mape_acc_sum,
                '_mape_acc_count':    _mape_acc_count,
                '_wape_abs_err':      _wape_abs_err,
                '_wmape_abs_err':     _wmape_abs_err,
            })

        rows.sort(key=lambda r: r['product'].lower())

        # ── Totales por mes ────────────────────────────────────────────────────
        month_totals = []
        for i, ym in enumerate(months):
            mfc = sum(r['cells'][i]['forecast']  for r in rows)
            mmo = sum(r['cells'][i]['mos']       for r in rows)
            mdl = sum(r['cells'][i]['delivered'] for r in rows)
            mso = sum(r['cells'][i]['so_demand'] for r in rows)
            month_totals.append({
                'month':     ym,
                'forecast':  round(mfc, 2),
                'mos':       round(mmo, 2),
                'delivered': round(mdl, 2),
                'so_demand': round(mso, 2),
            })

        # ── KPIs globales ──────────────────────────────────────────────────────
        total_fc   = sum(r['total_forecast']  for r in rows)
        total_mos  = sum(r['total_mos']       for r in rows)
        total_del  = sum(r['total_delivered'] for r in rows)
        total_so   = sum(r['total_so_demand'] for r in rows)

        # Producción de OFs para productos SIN línea de forecast (solo vendibles)
        no_fc_mo_pids = [pid for pid in mo_data if pid not in all_product_ids]
        mos_no_fc = 0.0
        if no_fc_mo_pids:
            sale_ok_pids = set(
                self.env['product.product'].browse(no_fc_mo_pids)
                .filtered(lambda p: p.sale_ok).ids
            )
            mos_no_fc = round(sum(
                sum(v.values()) for pid, v in mo_data.items()
                if pid not in all_product_ids and pid in sale_ok_pids
            ), 2)

        # Entregado para productos SIN línea de forecast (solo vendibles)
        delivered_no_fc = 0.0
        try:
            no_fc_del_domain = [
                ('state', '=', 'done'),
                ('picking_id.picking_type_id.code', '=', 'outgoing'),
                ('date', '>=', fields.Datetime.to_string(dt_from)),
                ('date', '<=', fields.Datetime.to_string(dt_to)),
                ('product_id.sale_ok', '=', True),
                ('company_id', '=', self.env.company.id),
            ]
            if all_product_ids_list:
                no_fc_del_domain.append(('product_id', 'not in', all_product_ids_list))
            groups = self.env['stock.move.line'].read_group(no_fc_del_domain, ['quantity:sum'], [])
            delivered_no_fc = round((groups[0]['quantity'] or 0.0) if groups else 0.0, 2)
        except Exception:
            delivered_no_fc = 0.0

        # Demanda de SOs en el período para productos SIN línea de forecast (solo vendibles).
        # El filtro sale_ok=True es consistente con mos_no_fc y delivered_no_fc; sin él
        # se inflaría el contador con líneas de productos internos o componentes.
        so_demand_no_fc = 0.0
        try:
            no_fc_domain = [
                ('order_id.state', 'in', ('sale', 'done')),
                ('order_id.date_order', '>=', fields.Datetime.to_string(dt_from)),
                ('order_id.date_order', '<=', fields.Datetime.to_string(dt_to)),
                ('product_id.sale_ok', '=', True),
                ('company_id', '=', self.env.company.id),
            ]
            if all_product_ids_list:
                no_fc_domain.append(('product_id', 'not in', all_product_ids_list))
            groups = self.env['sale.order.line'].read_group(no_fc_domain, ['product_uom_qty:sum'], [])
            so_demand_no_fc = round((groups[0]['product_uom_qty'] or 0.0) if groups else 0.0, 2)
        except Exception:
            so_demand_no_fc = 0.0

        coverage   = round(total_mos / total_fc * 100, 1) if total_fc > 0 else 0.0
        at_risk    = sum(1 for r in rows if r['total_forecast'] > 0 and r['total_pct'] < warning_pct)
        ovr_svc = round(total_del / total_so * 100, 1) if total_so > 0 else None
        _all_mape_sum   = sum(r['_mape_acc_sum']   for r in rows)
        _all_mape_count = sum(r['_mape_acc_count'] for r in rows)
        _all_wape_err   = sum(r['_wape_abs_err']   for r in rows)
        _all_wmape_err  = sum(r['_wmape_abs_err']  for r in rows)
        _total_fc_wmape = sum(r['total_forecast']  for r in rows if r['total_forecast'] > 0)
        total_actual = total_del if precision_source == 'delivery' else total_so
        acc_all = {
            'simple': round(total_actual / total_fc * 100, 1) if total_fc > 0 else None,
            'mape':   round(_all_mape_sum / _all_mape_count, 1) if _all_mape_count > 0 else None,
            'wape':   round(max(0.0, 100.0 - _all_wape_err / total_actual * 100), 1) if total_actual > 0 else None,
            'wmape':  round(max(0.0, 100.0 - _all_wmape_err / _total_fc_wmape * 100), 1) if _total_fc_wmape > 0 else None,
            'bias':   round((total_actual - total_fc) / total_fc * 100, 1) if total_fc > 0 else None,
        }
        ovr_acc = acc_all[acc_formula]

        # Limpiar campos internos antes de enviar al frontend
        _internal = ('_mape_acc_sum', '_mape_acc_count', '_wape_abs_err', '_wmape_abs_err')
        for r in rows:
            for k in _internal:
                r.pop(k, None)

        return {
            'kpis': {
                'total_forecast':       round(total_fc,  2),
                'total_so_demand':      round(total_so,  2),
                'total_mos':            round(total_mos, 2),
                'total_delivered':      round(total_del, 2),
                'gap':                  round(total_mos - total_fc, 2),
                'mos_gap_pct':          round((total_mos - total_fc) / total_fc * 100, 1) if total_fc > 0 else None,
                'demand_gap':           round(total_so - total_fc, 2),
                'demand_gap_pct':       round((total_so - total_fc) / total_fc * 100, 1) if total_fc > 0 else None,
                'coverage_pct':         coverage,
                'at_risk':              at_risk,
                'total_products':       len(rows),
                'overall_service_rate': ovr_svc,
                'overall_forecast_acc': ovr_acc,
                'acc_all':              acc_all,
                'so_demand_no_fc':      so_demand_no_fc,
                'mos_no_fc':            mos_no_fc,
                'delivered_no_fc':      delivered_no_fc,
            },
            'months':        months,
            'month_totals':  month_totals,
            'rows':          rows,
            'warning_pct':   warning_pct,
            'critical_pct':  critical_pct,
            'rotation_unit':    rotation_unit,
            'rotation_method':  rotation_method,
            'rotation_n_months': round(n_months, 2),
            'acc_formula':      acc_formula,
            'mo_states':        mo_states,
            'coverage_unit':            coverage_unit,
            'coverage_demand_source':   coverage_demand_source,
            'coverage_alerts_enabled':  coverage_alerts_enabled,
            'coverage_warn_days':       coverage_warn_days,
            'coverage_critical_days':   coverage_critical_days,
            'mo_coverage_show_pct':     mo_coverage_show_pct,
            'mo_coverage_denominator':  mo_coverage_denominator,
            'mo_coverage_color_scope':  mo_coverage_color_scope,
        }

    @api.model
    def get_product_mos_for_forecast(self, product_id, period_from, period_to, warehouse_ids=None):
        """
        Retorna las ÓFs de un producto para el acordeón de drill-down del widget de forecast.

        Usa la misma conversión UTC que get_forecast_dashboard_data para que los
        rangos de fecha sean consistentes entre la tabla principal y el drill-down.
        Las ÓFs se filtran por fecha de fin (date_finished) dentro del período y,
        opcionalmente, por ubicación destino de los almacenes seleccionados.

        :param product_id:   int — ID del product.product a consultar.
        :param period_from:  str — mes de inicio en formato 'YYYY-MM'.
        :param period_to:    str — mes de fin en formato 'YYYY-MM' (se toma el último día del mes).
        :param warehouse_ids: list[int] | None — IDs de almacenes para filtrar por location_dest_id.
                              Si es None o vacío, no se filtra por almacén.
        :returns: list[dict] — hasta 100 registros ordenados por date_finished asc, cada uno con:
            'id', 'name', 'state', 'state_label', 'product_qty', 'qty_produced',
            'uom', 'date_start', 'date_finished'.
        """
        from datetime import date as _date, datetime
        import calendar as _cal

        d_from = _date(int(period_from[:4]), int(period_from[5:7]), 1)
        d_to_y, d_to_m = int(period_to[:4]), int(period_to[5:7])
        last_day = _date(d_to_y, d_to_m, _cal.monthrange(d_to_y, d_to_m)[1])

        # Conversión a UTC para consistencia con get_forecast_dashboard_data.
        # Sin esta conversión, para un usuario en UTC-3 el acordeón y la tabla
        # principal mostrarían conjuntos de OFs distintos en los límites de mes.
        tz_name = self.env.context.get('tz') or self.env.user.tz or 'UTC'
        user_tz = pytz.timezone(tz_name)
        dt_from = user_tz.localize(
            datetime.combine(d_from, datetime.min.time())
        ).astimezone(pytz.utc).replace(tzinfo=None)
        dt_to = user_tz.localize(
            datetime.combine(last_day, datetime.max.time())
        ).astimezone(pytz.utc).replace(tzinfo=None)

        domain = [
            ('product_id', '=', product_id),
            ('state', 'not in', ['cancel']),
            ('date_finished', '>=', fields.Datetime.to_string(dt_from)),
            ('date_finished', '<=', fields.Datetime.to_string(dt_to)),
        ] + no_subcontract_domain(self.env)
        if warehouse_ids:
            wh_recs = self.env['stock.warehouse'].browse(warehouse_ids)
            loc_ids = wh_recs.mapped('lot_stock_id').ids
            if loc_ids:
                domain.append(('location_dest_id', 'in', loc_ids))

        mos = self.env['mrp.production'].search(domain, limit=100, order='date_finished asc')  # limit=100 para evitar payload excesivo en el acordeón
        state_labels = {
            'draft':     'Borrador',
            'confirmed': 'Confirmada',
            'progress':  'En progreso',
            'to_close':  'Por cerrar',
            'done':      'Hecha',
            'cancel':    'Cancelada',
        }
        return [{
            'id':           mo.id,
            'name':         mo.name,
            'state':        mo.state,
            'state_label':  state_labels.get(mo.state, mo.state),
            'product_qty':  round(mo.product_qty, 2),
            'qty_produced': round(mo.qty_produced, 2),
            'uom':          mo.product_uom_id.name if mo.product_uom_id else '',
            'date_start':   mo.date_start.strftime('%Y-%m-%d')    if mo.date_start    else None,
            'date_finished': mo.date_finished.strftime('%Y-%m-%d') if mo.date_finished else None,
        } for mo in mos]

    @api.model
    def get_forecast_export(self, period_from, period_to, warehouse_ids=None):
        """
        Genera un archivo Excel (.xlsx) con el resumen de forecast vs. ÓFs y retorna la URL de descarga.

        Internamente llama a get_forecast_dashboard_data para obtener los datos y luego
        construye un libro openpyxl con:
        - Fila 1: encabezados de meses (celdas combinadas Forecast + OFs).
        - Fila 2: sub-encabezados 'Forecast' / 'OFs' por mes.
        - Filas de datos: una por producto, coloreadas según cobertura (ok/warn/critical).
        - Fila de totales al final.
        El archivo se guarda como ir.attachment y se retorna su URL de descarga.

        :param period_from:  str — mes de inicio en formato 'YYYY-MM' o 'YYYY-MM-DD'.
        :param period_to:    str — mes de fin en formato 'YYYY-MM' o 'YYYY-MM-DD'.
        :param warehouse_ids: list[int] | None — se pasa directamente a get_forecast_dashboard_data.
        :returns: dict con clave 'url' (str) apuntando al endpoint /web/content/<id>?download=true,
                  o dict con clave 'error' (str) si openpyxl no está instalado.
        :raises: cualquier excepción de openpyxl o de escritura en ir.attachment se propaga.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return {'error': 'openpyxl no disponible'}
        import io, base64

        data = self.get_forecast_dashboard_data(period_from, period_to, warehouse_ids)
        months = data['months']
        rows   = data['rows']

        MONTHS_ES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
        def _label(ym):
            y, m = ym.split('-')
            return f"{MONTHS_ES[int(m)-1]} {y}"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Forecast'

        hdr_fill = PatternFill('solid', fgColor='1F497D')   # Azul corporativo para encabezados
        hdr_font = Font(bold=True, color='FFFFFF')
        ok_fill   = PatternFill('solid', fgColor='C6EFCE')  # Verde: cobertura >= 100 %
        warn_fill = PatternFill('solid', fgColor='FFEB9C')  # Amarillo: cobertura entre warning_pct y 100 %
        crit_fill = PatternFill('solid', fgColor='FFC7CE')  # Rojo: cobertura por debajo del umbral de alerta

        warning_pct  = data['warning_pct']

        # Fila 1: encabezados de meses (agrupados de a 2)
        col = 2
        ws.cell(1, 1, 'Artículo').font = hdr_font
        ws.cell(1, 1).fill = hdr_fill
        for ym in months:
            c1 = ws.cell(1, col, _label(ym))
            c1.font = hdr_font
            c1.fill = hdr_fill
            c1.alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
            col += 2
        ws.cell(1, col, 'Total Forecast').font = hdr_font
        ws.cell(1, col).fill = hdr_fill
        ws.cell(1, col + 1, 'Total OFs').font = hdr_font
        ws.cell(1, col + 1).fill = hdr_fill

        # Fila 2: sub-encabezados Forecast / OFs
        ws.cell(2, 1, 'Artículo').font = Font(bold=True)
        col = 2
        for _ in months:
            ws.cell(2, col, 'Forecast').font = Font(bold=True)
            ws.cell(2, col + 1, 'OFs').font = Font(bold=True)
            col += 2
        ws.cell(2, col, 'Forecast').font = Font(bold=True)
        ws.cell(2, col + 1, 'OFs').font = Font(bold=True)

        # Datos
        for r, row in enumerate(rows, start=3):
            ws.cell(r, 1, row['product'])
            col = 2
            for ci, ym in enumerate(months):
                cell = row['cells'][ci]
                fc_cell = ws.cell(r, col, cell['forecast'])
                mo_cell = ws.cell(r, col + 1, cell['mos'])
                if cell['forecast'] > 0:
                    fill = ok_fill if cell['pct'] >= 100 else (warn_fill if cell['pct'] >= warning_pct else crit_fill)
                    fc_cell.fill = fill
                    mo_cell.fill = fill
                col += 2
            ws.cell(r, col, row['total_forecast'])
            ws.cell(r, col + 1, row['total_mos'])

        # Fila de totales
        trow = len(rows) + 3
        ws.cell(trow, 1, 'TOTAL').font = Font(bold=True)
        col = 2
        for mt in data['month_totals']:
            ws.cell(trow, col, mt['forecast']).font = Font(bold=True)
            ws.cell(trow, col + 1, mt['mos']).font = Font(bold=True)
            col += 2
        ws.cell(trow, col, data['kpis']['total_forecast']).font = Font(bold=True)
        ws.cell(trow, col + 1, data['kpis']['total_mos']).font = Font(bold=True)

        ws.column_dimensions['A'].width = 30   # Ancho fijo para la columna de nombre de artículo
        for i in range(2, col + 2):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = 12  # Ancho estándar para columnas numéricas

        buf = io.BytesIO()
        wb.save(buf)
        content = base64.b64encode(buf.getvalue()).decode()

        attachment = self.env['ir.attachment'].create({
            'name': f'forecast_{period_from}_{period_to}.xlsx',
            'type': 'binary',
            'datas': content,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {'url': f'/web/content/{attachment.id}?download=true'}
